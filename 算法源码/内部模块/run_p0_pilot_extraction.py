# -*- coding: utf-8 -*-
from __future__ import annotations

import csv
import json
import math
import os
import re
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from rapidfuzz import fuzz


BASE_DIR = Path(__file__).resolve().parents[2]
SAMPLE_JSON = Path(os.environ.get(
    "SAMPLE_JSON_PATH",
    str(BASE_DIR / "算法源码" / "示例清单" / "示例样本清单.json"),
))
INDICATOR_JSON = Path(os.environ.get(
    "ESG_INDICATOR_JSON",
    str(BASE_DIR / "算法源码" / "配置" / "ESG指标体系.json"),
))
OUT_DIR = Path(os.environ.get("PILOT_OUT_DIR", str(BASE_DIR / "运行产物" / "候选抽取")))
PAGE_TEXT_DIR = OUT_DIR / "extracted_page_text"
OCR_CACHE_DIR = Path(os.environ.get("OCR_CACHE_DIR", str(BASE_DIR / "运行缓存" / "OCR")))
OCR_TEXT_DIR = OCR_CACHE_DIR / "ocr_pages"
OCR_JSON_DIR = OCR_CACHE_DIR / "ocr_page_json"
OCR_NATIVE_TEXT_MIN_CHARS = int(os.environ.get("OCR_NATIVE_TEXT_MIN_CHARS", "80"))
OCR_LAYOUT_LINE_MIN = int(os.environ.get("OCR_LAYOUT_LINE_MIN", "10"))
OCR_CID_RATIO_THRESHOLD = float(os.environ.get("OCR_CID_RATIO_THRESHOLD", "0.15"))
RUN_LABEL = os.environ.get("PILOT_RUN_LABEL", "5份试跑")

CSV_PATH = OUT_DIR / f"P0候选抽取结果_{RUN_LABEL}.csv"
JSON_PATH = OUT_DIR / f"P0候选抽取结果_{RUN_LABEL}.json"
XLSX_PATH = OUT_DIR / f"P0候选抽取结果_{RUN_LABEL}.xlsx"
LLM_JSONL_PATH = OUT_DIR / f"LLM校验提示队列_{RUN_LABEL}.jsonl"
REVIEW_XLSX_PATH = OUT_DIR / f"P0人工复核清单_{RUN_LABEL}.xlsx"
REPORT_PATH = OUT_DIR / "P0候选抽取试跑报告.md"

DEFAULT_PILOT_SAMPLE_IDS = ["GL001", "GL005", "GL008", "GL013", "GL030"]
EXTRACTOR_VERSION = "p0_candidate_pilot_v0.8c"

TABLE_CUES = [
    "关键绩效",
    "ESG绩效",
    "绩效表",
    "环境绩效",
    "社会绩效",
    "关键数据",
    "指标索引",
    "员工",
    "排放",
    "能源",
    "用水",
    "废弃物",
    "董事",
    "反腐",
]

INDEX_PAGE_CUES = ["目录", "索引表", "内容索引", "指标索引", "对应报告内容", "披露要求", "意见反馈"]
DATA_PAGE_CUES = [
    "ESG 数据表",
    "ESG数据表",
    "ESG绩效数据",
    "关键绩效",
    "环境绩效",
    "社会绩效",
    "绩效指标",
    "单位 数据",
    "类别 指标 单位",
]

STRUCTURED_VALUE_PAGE_CUES = [
    "可持续绩效总览",
    "关键绩效表",
    "关键绩效",
    "ESG 数据表",
    "ESG数据表",
    "ESG绩效数据",
    "环境绩效",
    "社会绩效",
    "治理绩效",
    "指标名称 单位",
    "类别 指标 单位",
]

UNIT_PATTERN = (
    r"万吨二氧化碳当量|吨二氧化碳当量|吨CO2e|kgCO2e|tCO2e|"
    r"万吨标准煤|万吨标煤|吨标准煤|吨标煤|万吨|吨|t|kg|"
    r"万千瓦时|千瓦时|kWh|MWh|GJ|百万千焦|"
    r"立方米|万立方米|m3|人次|人时|小时/人|小时|学时|"
    r"万元|亿元|元|人民币|人|名|次|场|件|起|宗|%|％"
)
NUMBER_PATTERN = re.compile(
    rf"(?<![A-Za-z])(?P<value>-?\d+(?:,\d{{3}})*(?:\.\d+)?)(?![A-Za-z])(?:\s*)(?P<unit>{UNIT_PATTERN})?",
    re.I,
)

CANDIDATE_COLUMNS = [
    "sample_id",
    "stock_code",
    "short_name",
    "report_type",
    "field_id",
    "dimension",
    "metric_name_cn",
    "metric_type",
    "value_type",
    "indicator_layer",
    "primary_indicator_id",
    "rating_role",
    "alternative_status_policy",
    "scoring_denominator_policy",
    "candidate_status",
    "candidate_disclosure_class",
    "candidate_rank",
    "evidence_type_candidate",
    "value_candidate",
    "unit_raw_candidate",
    "value_standardized_candidate",
    "unit_standardized_candidate",
    "value_status",
    "value_extraction_method",
    "source_page",
    "source_physical_page",
    "source_report_page_candidates",
    "source_text",
    "source_table_cell",
    "match_terms",
    "rule_score",
    "confidence_rule",
    "needs_llm_review",
    "review_reason",
    "recommended_next_status",
    "extractor_version",
    "pdf_path",
]

PDF_TEXT_FIXES = {
    "ÈË": "人",
    "％": "%",
}

EXTRA_TERMS = {
    "E_Q_001": ["范围一和范围二", "运营温室气体", "温室气体排放量", "碳排放总量", "公司法人边界排放", "法人边界排放"],
    "E_Q_002": ["范围一排放", "范围一温室气体排放", "温室气体排放（范围一）", "温室气体排放(范围一)", "直接排放", "发电设施碳排放", "发电设施 碳排放", "南湖厂区发电设施碳排放", "南湖厂区发电设施 碳排放"],
    "E_Q_003": ["范围二排放", "范围二温室气体排放", "温室气体排放（范围二）", "温室气体排放(范围二)", "间接排放", "外购电力对应的排放"],
    "E_Q_006": ["综合能耗", "总能耗", "能源消耗", "能耗总量", "能耗总量（等价值）", "能源消费总量", "能源总消耗量"],
    "E_Q_007": ["电力消耗", "耗电量", "用电量", "用电总量", "外购电", "外购电力", "办公用电消耗量", "办公用电"],
    "E_Q_009": ["用水量", "耗水量", "新鲜水", "水资源", "办公用水消耗量", "办公用水", "用水消耗量", "总用水", "取水总量", "取水量", "用水总量"],
    "E_Q_012": ["固体废物", "固体废弃物", "固体废弃物排放总量", "无害废弃物总量", "一般固废", "废物输出", "总废弃物量", "废弃物总产生量"],
    "E_Q_013": ["危险废物", "危险固体废物", "危废", "危险废弃物", "危险废弃物排放总量", "危险废物产生量", "有害废弃物", "危险废弃物产生量"],
    "E_Q_015": ["环保投资", "环保支出", "环境投入", "节能环保投入"],
    "E_T_001": ["应对气候变化", "双碳", "低碳转型", "绿色低碳", "碳达峰", "碳中和"],
    "E_T_002": ["双碳目标", "碳达峰", "碳中和", "减碳目标", "低碳目标"],
    "E_T_009": ["污染物排放", "废气治理", "废水治理", "节能降碳", "烟气", "排放管理"],
    "S_Q_001": ["员工总人数", "员工人数", "员工总数", "正式员工人数", "雇员总数"],
    "S_Q_004": ["培训总时长", "培训小时", "培训学时", "培训人时", "员工培训总时长", "员工培训总时数", "总培训小时", "总培训学时", "总时长", "培训总时数"],
    "S_Q_005": ["人均", "人均培训时长", "人均培训", "人均学时", "人均培训小时", "人均培训时间"],
    "S_Q_008": ["工伤率", "千人负伤率", "工伤事故率", "可记录工伤率", "LTIFR", "发生率", "伤亡事故率", "职业病害发生率", "安全事故发生率"],
    "S_Q_009": ["工亡", "死亡事故", "安全事故", "因工死亡", "未发生死亡", "职业健康相关环保事故", "安全环保事故", "伤亡事故", "员工伤亡"],
    "S_Q_017": ["公益", "捐赠", "慈善", "社会贡献", "志愿服务"],
    "G_Q_002": ["独立非执行董事", "独董", "独立董事"],
    "G_Q_003": ["独董占比", "独立董事占比", "独立董事人数占比", "董事会独立性"],
    "G_Q_009": ["廉洁培训", "反腐培训", "合规培训", "商业道德培训"],
    "G_Q_010": ["贪污诉讼", "腐败诉讼", "商业贿赂", "未发生贪污", "无贪污", "违法违规行为", "环境违法违规行为", "信访案件"],
    "G_T_002": ["董事会审议ESG", "董事会监督", "董事会可持续发展", "董事会ESG职责"],
    "G_T_004": ["合规体系", "内控体系", "内部控制", "审计监督"],
    "G_T_005": ["廉洁从业", "反舞弊", "反贿赂", "反贪污", "商业道德"],
}


def normalize_text(text: str) -> str:
    text = text or ""
    for old, new in PDF_TEXT_FIXES.items():
        text = text.replace(old, new)
    text = text.replace("，", ",").replace("．", ".")
    text = re.sub(r"\b[Yy](\d{1,3})\b", r"\1", text)
    return re.sub(r"\s+", " ", text).strip()


def unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        value = str(value or "").strip()
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def valid_report_page_number(value: int) -> bool:
    return 1 <= value <= 400 and not 1900 <= value <= 2100


def infer_report_page_candidates(physical_page: int, text: str) -> list[str]:
    text = normalize_text(text)
    if not text:
        return []
    candidates: list[str] = []
    head = text[:180]
    tail = text[-180:]

    leading_pair = re.match(r"^\s*(\d{1,3})\s+(\d{1,3})\b", head)
    if leading_pair:
        first = int(leading_pair.group(1))
        second = int(leading_pair.group(2))
        if valid_report_page_number(first) and valid_report_page_number(second) and abs(second - first) <= 2:
            candidates.extend([str(first), str(second)])
            doubled = physical_page * 2
            if abs(doubled - second) <= 2:
                candidates.extend([str(doubled - 1), str(doubled)])

    trailing = re.search(r"\b(\d{1,3})\s*$", tail)
    if trailing:
        page_num = int(trailing.group(1))
        if valid_report_page_number(page_num):
            candidates.append(str(page_num))

    return unique_preserve_order(candidates)


def source_page_reference(pdf_payload: dict[str, Any], physical_page: int | str) -> tuple[str, str, str]:
    physical = str(physical_page or "").strip()
    report_candidates: list[str] = []
    try:
        physical_int = int(str(physical_page))
    except (TypeError, ValueError):
        physical_int = 0
    if physical_int:
        for page in pdf_payload.get("pages", []):
            if int(page.get("page", 0)) == physical_int:
                report_candidates = list(page.get("report_page_candidates") or [])
                break
    source_pages = unique_preserve_order(([physical] if physical else []) + report_candidates)
    return ";".join(source_pages), physical, ";".join(report_candidates)


def load_ocr_payload_for_page(sample_id: str, physical_page: int) -> dict[str, Any] | None:
    json_path = OCR_JSON_DIR / f"{sample_id}_page_{physical_page:03d}_ocr.json"
    text_path = OCR_TEXT_DIR / f"{sample_id}_page_{physical_page:03d}_ocr.txt"
    if json_path.exists():
        try:
            payload = json.loads(json_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            payload = {}
        raw_text = str(payload.get("ocr_text") or "")
        text = normalize_text(raw_text)
        if text:
            line_payloads = payload.get("lines") or []
            ocr_lines = [
                str(item.get("text", "")).strip()
                for item in line_payloads
                if isinstance(item, dict) and str(item.get("text", "")).strip()
            ] or [line.strip() for line in raw_text.splitlines() if line.strip()]
            return {
                "text": text,
                "report_page_candidates": infer_ocr_report_page_candidates(physical_page, text),
                "ocr_lines": ocr_lines,
                "char_count": len(text),
                "ocr_line_count": payload.get("line_count", 0),
                "ocr_avg_score": payload.get("avg_score"),
                "ocr_cache_json": str(json_path),
            }
    if text_path.exists():
        raw_text = text_path.read_text(encoding="utf-8")
        text = normalize_text(raw_text)
        if text:
            return {
                "text": text,
                "report_page_candidates": infer_ocr_report_page_candidates(physical_page, text),
                "ocr_lines": [line.strip() for line in raw_text.splitlines() if line.strip()],
                "char_count": len(text),
                "ocr_line_count": 0,
                "ocr_avg_score": None,
                "ocr_cache_json": "",
            }
    return None


def native_text_is_garbled(text: str) -> bool:
    if not text:
        return False
    cid_count = len(re.findall(r"\(cid:\d+\)", text))
    if cid_count == 0:
        return False
    cid_ratio = (cid_count * 8) / max(1, len(text))
    return cid_ratio >= OCR_CID_RATIO_THRESHOLD


def infer_ocr_report_page_candidates(physical_page: int, text: str) -> list[str]:
    candidates = infer_report_page_candidates(physical_page, text)
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    numeric_tail = [int(line) for line in lines[-10:] if re.fullmatch(r"\d{1,3}", line)]
    if len(numeric_tail) >= 2 and 0 <= numeric_tail[-1] - numeric_tail[-2] <= 2:
        candidates.extend([str(numeric_tail[-2]), str(numeric_tail[-1])])
    elif numeric_tail:
        candidates.append(str(numeric_tail[-1]))
    return unique_preserve_order(candidates)


def split_aliases(*values: str) -> list[str]:
    terms: list[str] = []
    for value in values:
        if not value:
            continue
        for part in re.split(r"[;；,，/、|]+", value):
            part = part.strip()
            if len(part) >= 2 and part not in terms:
                terms.append(part)
    return terms


def split_units(*values: str) -> list[str]:
    units: list[str] = []
    for value in values:
        if not value:
            continue
        for part in re.split(r"[;；,，/、|]+", value):
            part = part.strip()
            if part and part not in units:
                units.append(part)
    return units


def terms_for_indicator(indicator: dict[str, Any]) -> list[str]:
    terms = split_aliases(
        indicator["metric_name_cn"],
        indicator.get("aliases_cn", ""),
        indicator.get("aliases_en", ""),
    )
    for term in EXTRA_TERMS.get(indicator["field_id"], []):
        if term not in terms:
            terms.append(term)
    return terms


def choose_sample_ids(samples: list[dict[str, Any]]) -> list[str]:
    env_ids = os.environ.get("PILOT_SAMPLE_IDS", "").strip()
    if env_ids:
        return [item.strip() for item in re.split(r"[,;，；\s]+", env_ids) if item.strip()]
    env_limit = os.environ.get("PILOT_SAMPLE_LIMIT", "").strip()
    if env_limit:
        try:
            limit = max(1, int(env_limit))
            return [item["sample_id"] for item in samples[:limit]]
        except ValueError:
            pass
    return DEFAULT_PILOT_SAMPLE_IDS


def load_inputs() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    samples = json.loads(SAMPLE_JSON.read_text(encoding="utf-8"))["samples"]
    indicators = json.loads(INDICATOR_JSON.read_text(encoding="utf-8"))["indicators"]
    sample_map = {item["sample_id"]: item for item in samples}
    selected_ids = choose_sample_ids(samples)
    missing = [sid for sid in selected_ids if sid not in sample_map]
    if missing:
        raise ValueError(f"Unknown sample_id(s): {', '.join(missing)}")
    selected = [sample_map[sid] for sid in selected_ids]
    p0_indicators = [item for item in indicators if item["extraction_priority"] == "P0"]
    return selected, p0_indicators


def extract_pdf(sample: dict[str, Any], global_clues: list[str]) -> dict[str, Any]:
    pdf_path = Path(sample["pdf_path"])
    pages: list[dict[str, Any]] = []
    table_rows: list[dict[str, Any]] = []
    start = time.time()
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_index, page in enumerate(pdf.pages, 1):
            native_text = normalize_text(page.extract_text() or "")
            native_cid_garbled = native_text_is_garbled(native_text)
            native_report_page_candidates = infer_report_page_candidates(page_index, native_text)
            ocr_payload = None
            candidate_ocr_payload = load_ocr_payload_for_page(sample["sample_id"], page_index)
            if candidate_ocr_payload:
                ocr_text = str(candidate_ocr_payload.get("text") or "")
                ocr_char_count = int(candidate_ocr_payload.get("char_count", 0))
                ocr_line_count = int(candidate_ocr_payload.get("ocr_line_count", 0))
                ocr_has_clue = any(term and term in ocr_text for term in global_clues)
                ocr_is_substantially_better = (
                    (len(native_text) < OCR_NATIVE_TEXT_MIN_CHARS or native_cid_garbled)
                    and ocr_char_count > max(len(native_text) + 40, len(native_text) * 2)
                )
                ocr_repairs_garbled_text = native_cid_garbled and ocr_char_count >= 20
                ocr_layout_is_better = (
                    ocr_has_clue
                    and ocr_line_count >= OCR_LAYOUT_LINE_MIN
                    and ocr_char_count >= max(20, int(len(native_text) * 0.75))
                )
                if ocr_is_substantially_better or ocr_repairs_garbled_text or ocr_layout_is_better:
                    ocr_payload = candidate_ocr_payload
            if (native_text and not native_cid_garbled) or not ocr_payload:
                pages.append(
                    {
                        "page": page_index,
                        "report_page_candidates": native_report_page_candidates,
                        "text": native_text,
                        "char_count": len(native_text),
                        "text_source": "native_text" if native_text else "",
                        "native_char_count": len(native_text),
                    }
                )
            if ocr_payload:
                ocr_report_page_candidates = unique_preserve_order(
                    list(native_report_page_candidates) + list(ocr_payload.get("report_page_candidates") or [])
                )
                pages.append(
                    {
                        "page": page_index,
                        "report_page_candidates": ocr_report_page_candidates,
                        "text": ocr_payload["text"],
                        "char_count": len(ocr_payload["text"]),
                        "text_source": "ocr_text",
                        "native_char_count": len(native_text),
                        "ocr_line_count": ocr_payload.get("ocr_line_count", 0),
                        "ocr_avg_score": ocr_payload.get("ocr_avg_score"),
                        "ocr_cache_json": ocr_payload.get("ocr_cache_json", ""),
                        "ocr_lines": ocr_payload.get("ocr_lines", []),
                    }
                )
            should_try_tables = bool(native_text) and any(term and term in native_text for term in global_clues)
            if should_try_tables:
                try:
                    tables = page.extract_tables() or []
                except Exception as exc:
                    table_rows.append(
                        {
                            "page": page_index,
                            "row_text": "",
                            "table_index": -1,
                            "row_index": -1,
                            "extract_error": f"{type(exc).__name__}: {exc}",
                        }
                    )
                    continue
                for table_index, table in enumerate(tables, 1):
                    normalized_rows = [
                        [normalize_text(str(cell or "")) for cell in row]
                        for row in (table or [])
                    ]
                    header_candidates: list[str] = []
                    for row in normalized_rows[:4]:
                        row_text_for_header = " | ".join(cell for cell in row if cell)
                        if row_text_for_header:
                            header_candidates.append(row_text_for_header)
                    for row in normalized_rows:
                        row_text_for_header = " | ".join(cell for cell in row if cell)
                        if any(term in row_text_for_header for term in ["指标", "单位", "2023", "2024", "2025"]):
                            header_candidates.append(row_text_for_header)
                    header_context = " || ".join(dict.fromkeys(header_candidates[:6]))
                    for row_index, cells in enumerate(normalized_rows, 1):
                        row_text = " | ".join(cell for cell in cells if cell)
                        if len(row_text) >= 5:
                            table_rows.append(
                                {
                                    "page": page_index,
                                    "report_page_candidates": native_report_page_candidates,
                                    "row_text": row_text,
                                    "cells": cells,
                                    "header_context": header_context,
                                    "table_index": table_index,
                                    "row_index": row_index,
                                    "extract_error": "",
                                }
                            )
    payload = {
        "sample_id": sample["sample_id"],
        "stock_code": sample["stock_code"],
        "short_name": sample["short_name"],
        "pdf_path": str(pdf_path),
        "page_count": len(pages),
        "pages": pages,
        "table_rows": table_rows,
        "elapsed_sec": round(time.time() - start, 3),
    }
    (PAGE_TEXT_DIR / f"{sample['sample_id']}_{sample['stock_code']}_page_text.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return payload


def page_score(text: str, terms: list[str]) -> tuple[float, list[str]]:
    if not text:
        return 0.0, []
    hits: list[str] = []
    score = 0.0
    for term in terms:
        exact_score = 16 + min(len(term), 12)
        if term in text:
            hits.append(term)
            score += exact_score
        elif len(term) >= 4:
            ratio = fuzz.partial_ratio(term, text)
            if ratio >= 85:
                hits.append(f"{term}~{ratio}")
                score += exact_score * 0.6
    return score, hits


def evidence_page_adjustment(text: str, indicator: dict[str, Any]) -> float:
    if not text:
        return 0.0
    is_quantitative = indicator.get("metric_type") == "quantitative"
    adjustment = 0.0
    if any(cue in text for cue in DATA_PAGE_CUES):
        adjustment += 28 if is_quantitative else 8
    if any(cue in text for cue in INDEX_PAGE_CUES):
        adjustment -= 70 if is_quantitative else 18
    return adjustment


def evidence_structure_bonus(text: str, indicator: dict[str, Any]) -> float:
    if not text:
        return 0.0
    bonus = 0.0
    metric_name = indicator.get("metric_name_cn", "")
    if metric_name and metric_name in text:
        bonus += 24
    if indicator.get("metric_type") == "quantitative":
        if any(cue in text for cue in STRUCTURED_VALUE_PAGE_CUES):
            bonus += 24
        if "指标名称" in text and "单位" in text and "2025" in text:
            bonus += 18
        if re.search(r"2023\s*年.{0,20}2024\s*年.{0,20}2025\s*年", text):
            bonus += 16
        if "目录" in text[:180] and not any(cue in text for cue in STRUCTURED_VALUE_PAGE_CUES):
            bonus -= 30
    if "指标索引" in text and indicator.get("metric_type") == "quantitative" and "可持续绩效总览" not in text:
        bonus -= 8
    return bonus


def snippet_around(text: str, terms: list[str], width: int = 260) -> str:
    if not text:
        return ""
    positions = [(text.find(term), len(term)) for term in terms if term and term in text]
    positions = [item for item in positions if item[0] >= 0]
    if positions:
        pos = sorted(positions, key=lambda item: (-item[1], item[0]))[0][0]
    else:
        pos = -1
    if pos < 0:
        return text[: width * 2]
    start = max(0, pos - width)
    end = min(len(text), pos + width)
    return text[start:end]


def accepted_units_for(indicator: dict[str, Any]) -> list[str]:
    units = split_units(indicator.get("unit_normalized", ""), indicator.get("units_accepted_raw", ""))
    expanded: list[str] = []
    for unit in units:
        expanded.append(unit)
        if unit == "%":
            expanded.append("％")
        if unit == "person":
            expanded.extend(["人", "名"])
        if unit == "case":
            expanded.extend(["件", "起", "宗", "次"])
        if unit == "CNY":
            expanded.extend(["元", "万元", "亿元", "人民币"])
        if unit == "t":
            expanded.extend(["吨", "kg", "千克"])
        if unit == "tCO2e":
            expanded.extend([
                "tCO2e",
                "吨CO2e",
                "kgCO2e",
                "千克CO2e",
                "吨二氧化碳当量",
                "万吨二氧化碳当量",
                "吨CO2e/万人次",
                "千克CO2e/万元",
                "万吨",
            ])
        if unit == "m3":
            expanded.extend(["m3", "立方米", "万立方米"])
        if unit in ["MWh", "kWh"]:
            expanded.extend(["千瓦时", "万千瓦时", "MWh", "kWh"])
        if unit == "MWh":
            expanded.extend(["吨标煤", "吨标准煤", "万吨标准煤", "万吨标煤"])
    return sorted(set(expanded), key=len, reverse=True)


def default_unit_for_zero(indicator: dict[str, Any]) -> str:
    if indicator["field_id"] == "S_Q_009":
        return "人"
    if indicator["field_id"] == "G_Q_010":
        return "件"
    return indicator.get("unit_normalized", "")


def explicit_zero_value(text: str, indicator: dict[str, Any]) -> tuple[str, str] | None:
    if indicator["field_id"] not in {"S_Q_009", "G_Q_010"}:
        return None
    if indicator["field_id"] == "S_Q_009":
        if re.search(r"(未发生|无|零).{0,12}(工亡|死亡|致命|安全事故)", text):
            return "0", default_unit_for_zero(indicator)
        if re.search(r"(未发生|无|零).{0,20}(员工|职业健康|安全|工伤).{0,12}(事故|事件|死亡|工亡)", text):
            return "0", default_unit_for_zero(indicator)
    if indicator["field_id"] == "G_Q_010":
        if re.search(r"(未发生|无|零).{0,16}(腐败|贪污|贿赂|诉讼|案件|违规)", text):
            return "0", default_unit_for_zero(indicator)
        if re.search(r"(未发生|无|零).{0,20}(违法违规|违规行为|信访案件|处罚|刑事责任)", text):
            return "0", default_unit_for_zero(indicator)
    return None


def infer_unit_near(text: str, start: int, end: int, indicator: dict[str, Any]) -> str:
    accepted_units = accepted_units_for(indicator)
    before = text[max(0, start - 60) : start]
    after = text[end : min(len(text), end + 20)]
    near = before + " " + after
    for unit in accepted_units:
        if unit and unit in near:
            return unit
    matches = list(re.finditer(UNIT_PATTERN, near, re.I))
    return matches[-1].group(0) if matches else ""


def number_match_is_noise(text: str, match: re.Match[str]) -> bool:
    unit = match.group("unit") or ""
    value_text = match.group("value") or ""
    start, end = match.span()
    left = text[max(0, start - 20) : start]
    right = text[end : min(len(text), end + 12)]
    near = left + value_text + right

    # ── v2.3: 化学式/公式下标检测（优先于单位检查）──
    # 即使有单位也要检查，因为 "COe 2 吨" 中的 "2" 也是噪声
    try:
        numeric_val = float(value_text.replace(",", ""))
    except ValueError:
        numeric_val = 0
    if numeric_val <= 99:
        # v2.3: 宽窗口化学式检测（前120字符）
        # 覆盖长文本中 "COe ... (吨 2) ... (吨 2)" 等远距离下标模式
        wide_window = text[max(0, start - 120) : min(len(text), end + 30)]
        if re.search(r"(CO2?e?|NOx?|SO[x2]|CH4|N2O|HFC|PFC|SF6|PM\d*|CO?2?\s*e?)", wide_window, re.I):
            return True
        # 全文化学标记 + 1-9单数字无单位 → 下标噪声（保守策略）
        if numeric_val <= 9 and not unit:
            if re.search(r"\b(CO2|COe|NOx|SO[x2]|CH4|N2O|HFC|PFC|SF6|PM\d+)\b", text, re.I):
                return True
        # 小整数紧跟大写字母缩写
        if re.search(r"[A-Z]{2,6}\b.{0,10}\b\d{1,2}\b", wide_window):
            return True

    if unit:
        return False
    if end < len(text) and text[end : end + 1] == "." and text[end + 1 : end + 2].isdigit():
        return True
    if start > 0 and text[start - 1 : start] == "." and text[start - 2 : start - 1].isdigit():
        return True
    if re.search(r"(范围|范畴|类别|第)\s*[一二三四五六七八九十0-9]+", near):
        return True
    if re.search(r"(ISO|GB|HJ|T/|No\.?|编号|标准)\s*[-A-Za-z0-9]*$", left + value_text, re.I):
        return True
    if re.search(r"[0-9]+(?:\.[0-9]+){2,}\s*$", left + value_text):
        return True
    # ── v2.3: 化学式下标窄窗口检测（无单位时更严格）──
    if re.search(r"(CO2?e?|NOx?|SO[x2]|CH4|N2O)\s*\d{1,2}\b", left + value_text, re.I):
        return True
    return False


def non_year_number_matches(text: str) -> list[re.Match[str]]:
    matches: list[re.Match[str]] = []
    for match in NUMBER_PATTERN.finditer(text):
        value = float((match.group("value") or "0").replace(",", ""))
        unit = match.group("unit") or ""
        if 1900 <= value <= 2100 and not unit:
            continue
        if number_match_is_noise(text, match):
            continue
        matches.append(match)
    return matches


def year_sequence(text: str) -> list[int]:
    years: list[int] = []
    for match in re.finditer(r"20(?:23|24|25)", text):
        year = int(re.sub(r"\D", "", match.group(0))[:4])
        if year not in years:
            years.append(year)
    # Table values follow the visual header order, which is often descending
    # (2025, 2024, 2023). Sorting breaks the year-to-column binding.
    return years


def year_aware_number(text: str, indicator: dict[str, Any]) -> tuple[str, str] | None:
    if indicator["metric_type"] != "quantitative":
        return None
    terms = terms_for_indicator(indicator)
    term_positions = [
        (text.find(term), term)
        for term in sorted(terms, key=len, reverse=True)
        if term and text.find(term) >= 0
    ]
    if not term_positions:
        return None
    for pos, _term in term_positions:
        context = text[max(0, pos - 360) : min(len(text), pos + 620)]
        after = text[pos : min(len(text), pos + 420)]

        # ── v2.4: Scope边界限制 ──
        # 对于范围一/二/三排放指标，限定搜索窗口不超过下一个scope标记
        field_id = indicator.get("field_id", "")
        scope_boundaries = {
            "E_Q_002": ["范围二", "范围三", "温室气体排放总量", "范围一和范围二"],
            "E_Q_003": ["范围三", "范围一和范围二", "温室气体排放总量"],
            "E_Q_004": ["温室气体排放总量"],
        }
        if field_id in scope_boundaries:
            for boundary_term in scope_boundaries[field_id]:
                boundary_pos = after.find(boundary_term)
                if 15 <= boundary_pos < len(after):
                    after = after[:boundary_pos]
                    break

        matches_after = non_year_number_matches(after)
        if not matches_after:
            continue
        number_token = r"-?\d+(?:,\d{3})*(?:\.\d+)?"
        unit_sequence = re.search(
            rf"(?P<unit>{UNIT_PATTERN})\s+(?P<values>{number_token}(?:\s+{number_token}){{0,2}})",
            after[:160],
            re.I,
        )
        if unit_sequence:
            unit_before = unit_sequence.group("unit")
            values = re.findall(number_token, unit_sequence.group("values"))
            if values and unit_is_compatible(unit_before, indicator):
                candidate_value = values[-1].replace(",", "")
                # ── v2.3: 噪声检查（防止 "COe 吨 2" 中的下标被当作值）──
                # 在 unit_sequence 匹配位置附近检查化学式标记
                try:
                    candidate_numeric = float(candidate_value)
                except ValueError:
                    candidate_numeric = 0
                if candidate_numeric <= 99:
                    us_start = unit_sequence.start()
                    us_window = after[max(0, us_start - 40):min(len(after), unit_sequence.end() + 30)]
                    # 同时检查 after 全文是否有化学标记（因为 COe 可能在 term 位置之前）
                    after_has_chem = bool(re.search(
                        r"\b(CO2|COe|NOx|SO[x2]|CH4|N2O|HFC|PFC|SF6|PM\d+)\b",
                        after, re.I
                    ))
                    if after_has_chem or re.search(
                        r"(CO2?e?|NOx?|SO[x2]|CH4|N2O|HFC|PFC|SF6|PM\d*)",
                        us_window, re.I
                    ):
                        # 疑似化学下标，跳过 unit_sequence 快速路径，
                        # 继续使用 matches_after 的多值逻辑
                        pass
                    else:
                        return candidate_value, unit_before
                else:
                    return candidate_value, unit_before
        first = matches_after[0]
        first_direct_unit = first.group("unit") or ""
        if first.start() <= 40 and first_direct_unit and unit_is_compatible(first_direct_unit, indicator):
            return (first.group("value") or "").replace(",", ""), first_direct_unit
        year_context = context[:420]
        years = year_sequence(year_context + " " + after[:160])
        has_multi_year = len(years) >= 2
        if 2025 in years and len(matches_after) >= len(years):
            selected = matches_after[years.index(2025)]
        elif has_multi_year and len(matches_after) >= len(years):
            selected = matches_after[-1]
        else:
            selected = matches_after[0]
        value = (selected.group("value") or "").replace(",", "")
        unit = selected.group("unit") or infer_unit_near(after, selected.start(), selected.end(), indicator)
        if unit and not unit_is_compatible(unit, indicator):
            continue
        if indicator["value_type"] != "percentage" and unit in ["%", "％"]:
            continue
        return value, unit
    return None


def direct_labeled_number(text: str, indicator: dict[str, Any]) -> tuple[str, str] | None:
    if indicator["metric_type"] != "quantitative":
        return None
    for term in sorted(terms_for_indicator(indicator), key=len, reverse=True):
        pos = text.find(term)
        if pos < 0:
            continue
        segment = text[pos : min(len(text), pos + 180)]

        # ── v2.4: Scope边界限制 ──
        field_id = indicator.get("field_id", "")
        scope_boundaries = {
            "E_Q_002": ["范围二", "范围三", "温室气体排放总量"],
            "E_Q_003": ["范围三", "温室气体排放总量"],
            "E_Q_004": ["温室气体排放总量"],
        }
        if field_id in scope_boundaries:
            for boundary_term in scope_boundaries[field_id]:
                boundary_pos = segment.find(boundary_term)
                if 10 <= boundary_pos < len(segment):
                    segment = segment[:boundary_pos]
                    break

        matches = non_year_number_matches(segment)
        scored: list[tuple[float, str, str]] = []
        for match in matches:
            value = (match.group("value") or "").replace(",", "")
            unit = match.group("unit") or infer_unit_near(segment, match.start(), match.end(), indicator)
            numeric_value = float(value or "0")
            if unit and not unit_is_compatible(unit, indicator):
                continue
            if indicator["value_type"] != "percentage" and unit in ["%", "％"]:
                continue
            score = 120 - match.start()
            if unit:
                score += 50
            if numeric_value == 100 and unit == "%" and indicator["value_type"] != "percentage":
                score -= 80
            scored.append((score, value, unit))
        if scored:
            scored.sort(key=lambda item: item[0], reverse=True)
            _, value, unit = scored[0]
            return value, unit
    return None


def extract_number(text: str, indicator: dict[str, Any]) -> tuple[str, str]:
    if not text:
        return "", ""
    if indicator["metric_type"] != "quantitative":
        return "", ""
    zero_value = explicit_zero_value(text, indicator)
    if zero_value:
        return zero_value
    # ── v2.4: G_Q_001 董事会总数特殊提取 ──
    # "董事会由 X 名董事组成" 模式明确表示总数
    if indicator.get("field_id") == "G_Q_001":
        # OCR文本中 "董事会由 12 名董事组成" 常被拆分
        # 策略1: "董事会由 ... X ... 名董事"
        board_match = re.search(r"董事会由.{0,40}?(\d+)\s*名", text)
        if board_match:
            val = int(board_match.group(1))
            if 3 <= val <= 30:  # 合理董事会规模
                return board_match.group(1), "人"
        # 策略2: 在包含"董事会"的段落中，找5-20之间的数字（合理规模）
        if "董事会" in text or "董事" in text:
            for m in re.finditer(r"\b(\d+)\b", text):
                val = int(m.group(1))
                if 7 <= val <= 20:  # 典型董事会规模
                    # 检查附近是否有 "董事" 相关词
                    nearby = text[max(0,m.start()-30):min(len(text),m.end()+30)]
                    if re.search(r"董事|董事会", nearby):
                        return m.group(1), "人"
    year_value = year_aware_number(text, indicator)
    if year_value:
        return year_value
    direct_value = direct_labeled_number(text, indicator)
    if direct_value:
        return direct_value
    matches = non_year_number_matches(text)
    if not matches:
        return "", ""
    terms = terms_for_indicator(indicator)
    term_positions = [text.find(term) for term in terms if term and text.find(term) >= 0]
    accepted_units = accepted_units_for(indicator)
    scored: list[tuple[float, re.Match[str], str]] = []
    for match in matches:
        unit = match.group("unit") or infer_unit_near(text, match.start(), match.end(), indicator)
        value = match.group("value") or ""
        numeric_value = float(value.replace(",", ""))
        score = 0.0
        if 1900 <= numeric_value <= 2100 and not unit:
            score -= 30
        if term_positions:
            best_distance = min(abs(match.start() - pos) for pos in term_positions)
            score += max(0, 80 - best_distance) / 2
            if any(pos <= match.start() for pos in term_positions):
                score += 12
        if unit:
            score += 5
        if unit and unit in accepted_units:
            score += 24
        elif unit and not unit_is_compatible(unit, indicator):
            continue
        if indicator["value_type"] == "percentage" and unit in ["%", "％"]:
            score += 15
        if indicator["value_type"] in ["integer"] and numeric_value.is_integer():
            score += 3
        if indicator["value_type"] == "currency" and unit in ["元", "万元", "亿元", "人民币"]:
            score += 15
        if indicator["unit_normalized"] and indicator["unit_normalized"] in ["tCO2e", "MWh", "kWh", "m3", "t"]:
            score += 6 if unit else 0
        scored.append((score, match, unit))
    if not scored:
        return "", ""
    scored.sort(key=lambda item: item[0], reverse=True)
    best_score, best, best_unit = scored[0]
    best_value = float((best.group("value") or "0").replace(",", ""))
    if 1900 <= best_value <= 2100 and not best_unit:
        return "", ""
    if best_score <= 3 and not best_unit:
        return "", ""
    # ── v2.3: 小整数噪声过滤 ──
    # 1-9 的无单位整数，如果离所有指标术语都 >20 字符，视为噪声
    best_numeric = float(best_value)
    if 1 <= best_numeric <= 9 and not best_unit:
        terms = terms_for_indicator(indicator)
        terms_nearby = False
        for term in terms:
            if term:
                pos = text.find(term)
                if pos >= 0 and abs(best.start() - pos) < 20:
                    terms_nearby = True
                    break
        if not terms_nearby:
            return "", ""
    return (best.group("value") or "").replace(",", ""), best_unit


def is_ocr_metric_label_line(line: str) -> bool:
    text = normalize_text(line)
    if not text or len(text) > 42:
        return False
    number_matches = non_year_number_matches(text)
    if number_matches:
        has_unit = any(match.group("unit") for match in number_matches)
        if has_unit:
            return False
        if not re.search(r"(占比|比例|人数|员工|董事|会议|岁|以下|以上)", text):
            return False
    if re.fullmatch(r"[A-Za-z0-9./:：、（）() -]+", text):
        return False
    if text in {"报告导读", "公司概况", "年度数据", "环境", "社会", "治理", "经济", "安全", "员工"}:
        return False
    return bool(re.search(r"[\u4e00-\u9fff]", text))


def ordered_number_unit_matches(
    text: str,
    indicator: dict[str, Any] | None = None,
    filter_units: bool = True,
) -> list[tuple[str, str]]:
    normalized = normalize_text(text)
    results: list[tuple[str, str]] = []
    for match in non_year_number_matches(normalized):
        value = (match.group("value") or "").replace(",", "")
        unit = match.group("unit") or ""
        if filter_units and indicator and unit and not unit_is_compatible(unit, indicator):
            continue
        if filter_units and indicator and indicator.get("value_type") != "percentage" and unit in ["%", "％"]:
            continue
        results.append((value, unit))
    return results


def extract_ocr_year_column_value(cleaned_lines: list[str], index: int, indicator: dict[str, Any]) -> tuple[str, str, str] | None:
    header = " ".join(cleaned_lines[:18])
    if "2024年" not in header or "2025年" not in header:
        return None
    for unit_index in range(index + 1, min(len(cleaned_lines), index + 5)):
        unit_line = cleaned_lines[unit_index]
        if ordered_number_unit_matches(unit_line, filter_units=False):
            continue
        if not unit_is_compatible(unit_line, indicator):
            continue
        values: list[str] = []
        for value_index in range(unit_index + 1, min(len(cleaned_lines), unit_index + 7)):
            value_line = cleaned_lines[value_index]
            if is_ocr_metric_label_line(value_line) and values:
                break
            for value, unit in ordered_number_unit_matches(value_line, filter_units=False):
                if unit and not unit_is_compatible(unit, indicator):
                    continue
                values.append(value)
            if len(values) >= 2:
                if (
                    len(values) == 2
                    and values[0].isdigit()
                    and values[1].isdigit()
                    and 1 <= int(values[0]) <= 999
                    and 0 <= int(values[1]) <= 99
                ):
                    return f"{values[0]}.{values[1].zfill(2)}", unit_line, "ocr_year_column_decimal_split_rule"
                return values[1], unit_line, "ocr_year_column_rule"
    return None


def _premerge_ocr_split_numbers(lines: list[str]) -> list[str]:
    """v2.2: 预合并 OCR 拆分的数字。
    处理两种常见 OCR 拆分:
    1. 逗号拆分: '15' + '192.84 万千瓦时' → 合并为 '15192.84 万千瓦时'
    2. 小数点拆分: '56' + '89' → '56.89' (v0.8c 已有，此处扩展)
    """
    merged = []
    i = 0
    while i < len(lines):
        line = lines[i]
        # 检查当前行是否为短数字（1-3位整数），下一行以数字开头且含单位
        if re.fullmatch(r"\d{1,3}", line) and i + 1 < len(lines):
            next_line = lines[i + 1]
            next_num = re.match(r"(\d{1,3}(?:\.\d+)?)\s*(.*)", next_line)
            if next_num:
                # 逗号拆分: "15" + "192.84 万千瓦时" → "15192.84 万千瓦时"
                combined_val = line + next_num.group(1)
                combined_unit = next_num.group(2) or ""
                # 过滤：合并后的值应该是合理的大数
                try:
                    if float(combined_val) > float(line) * 10:
                        merged.append(f"{combined_val} {combined_unit}".strip())
                        i += 2
                        continue
                except ValueError:
                    pass
            # 小数点拆分: "56" + "89" → "56.89" (v0.8c rule)
            if re.fullmatch(r"\d{1,2}", next_line):
                merged.append(f"{line}.{next_line}")
                i += 2
                continue
        merged.append(line)
        i += 1
    return merged


def _ocr_line_specificity(line: str, indicator: dict[str, Any]) -> float:
    """v2.3: 计算 OCR 行对特定指标的特异性得分。

    用于过滤 extract_number_from_ocr_lines 中的候选行，
    防止 EXTRA_TERMS 通用词导致跨指标误匹配。

    主名称命中 → 100，别名命中 → 70，
    EXTRA_TERMS 命中 → 40，无命中 → 0。
    """
    primary_terms = split_aliases(indicator.get("metric_name_cn", ""))
    alias_terms = split_aliases(
        indicator.get("aliases_cn", ""), indicator.get("aliases_en", "")
    )
    extra_terms = EXTRA_TERMS.get(indicator["field_id"], [])

    # 1. 主名称精确命中
    for pt in primary_terms:
        if pt and len(pt) >= 4 and pt in line:
            return 100.0
        if pt and len(pt) >= 4 and fuzz.partial_ratio(pt, line) >= 95:
            return 85.0

    # 2. 别名命中
    for at in alias_terms:
        if at and len(at) >= 3 and at in line:
            return 70.0
        if at and len(at) >= 4 and fuzz.partial_ratio(at, line) >= 92:
            return 55.0

    # 3. EXTRA_TERMS 命中
    for et in extra_terms:
        if et and len(et) >= 2 and et in line:
            return 40.0

    return 0.0


def extract_number_from_ocr_lines(lines: list[str], indicator: dict[str, Any]) -> tuple[str, str, str] | None:
    if indicator["metric_type"] != "quantitative" or not lines:
        return None
    terms = sorted(terms_for_indicator(indicator), key=len, reverse=True)
    cleaned_lines = [normalize_text(line) for line in lines if normalize_text(line)]
    # ── v2.2: 预合并 OCR 拆分数字 ──
    cleaned_lines = _premerge_ocr_split_numbers(cleaned_lines)
    full_text = " ".join(cleaned_lines)
    zero_value = explicit_zero_value(full_text, indicator)
    if zero_value:
        value, unit = zero_value
        return value, unit, "ocr_explicit_zero_rule"

    def indices_for(search_terms: list[str]) -> list[int]:
        return [
            index
            for index, line in enumerate(cleaned_lines)
            if any(term and term in line for term in search_terms if term)
        ]

    primary_terms = split_aliases(indicator.get("metric_name_cn", ""))
    alias_terms = split_aliases(indicator.get("aliases_cn", ""), indicator.get("aliases_en", ""))
    extra_terms = list(EXTRA_TERMS.get(indicator["field_id"], []))
    candidate_indices = (
        indices_for(primary_terms)
        or indices_for(alias_terms)
        or indices_for(extra_terms)
        or indices_for(terms)
    )
    preferred_ocr_terms = {
        "E_Q_006": ["综合能源消耗量", "综合能源消耗", "综合能耗", "总能耗", "能源消耗总量", "能源总消耗量"],
        "E_Q_007": ["外购电力", "外购电量", "用电量", "电力消耗", "耗电量"],
        "E_Q_009": ["总用水量", "取水量", "用水量", "水", "新鲜水用量"],
        "E_Q_012": ["废弃物产生总量", "废弃物总量", "固体废弃物产生总量", "一般固体废弃物排放总量", "废弃物排放量", "总废弃物量"],
        "E_Q_013": ["危险废弃物产生量", "危险废物产生量", "危险废弃物", "危废产生量", "有害废弃物"],
        "E_Q_014": ["一般废弃物产生量", "非危险废弃物", "无害废弃物产生量", "无害废弃物量"],
        "S_Q_004": ["培训总时长", "培训总时数", "总时长", "总培训小时", "总培训学时", "员工培训总时长"],
        "S_Q_005": ["人均培训时长", "人均培训", "人均学时", "人均培训小时", "平均培训时间", "人均时长"],
        "S_Q_008": ["工伤率", "发生率", "伤亡事故率", "千人负伤率", "可记录工伤率"],
        "S_Q_009": ["工亡人数", "因工死亡人数", "死亡事故", "工亡", "工伤死亡", "因工亡故"],
        "G_Q_002": ["独立非执行董事人数", "独立董事人数", "独董人数", "独立非执行董事"],
        "G_Q_003": ["独立董事占比", "独董占比", "董事会独立性", "独立董事比例"],
        "G_Q_009": ["廉洁培训", "反腐培训", "商业道德培训", "合规培训人次", "反腐败培训"],
        "G_Q_010": ["贪污诉讼", "腐败案件", "商业贿赂案件", "腐败诉讼", "违规案件"],
    }
    preferred_indices: list[int] = []
    for preferred_term in preferred_ocr_terms.get(indicator["field_id"], []):
        for index, line in enumerate(cleaned_lines):
            if len(preferred_term) == 1:
                if line == preferred_term:
                    preferred_indices.append(index)
            elif preferred_term in line:
                preferred_indices.append(index)
    if preferred_indices:
        seen_indices: set[int] = set()
        candidate_indices = [
            index
            for index in preferred_indices + candidate_indices
            if not (index in seen_indices or seen_indices.add(index))
        ]

    # ── v2.3: 特异性过滤 ──
    # 仅保留对当前指标有足够特异性的行，防止 EXTRA_TERMS 通用词跨指标误匹配
    if len(candidate_indices) > 1:
        scored_indices = [
            (index, _ocr_line_specificity(cleaned_lines[index], indicator))
            for index in candidate_indices
        ]
        # 过滤掉特异性 < 40 的行（除非是唯一候选）
        filtered = [(idx, sc) for idx, sc in scored_indices if sc >= 40.0]
        if filtered:
            # 按特异性降序排列
            filtered.sort(key=lambda x: x[1], reverse=True)
            candidate_indices = [idx for idx, _ in filtered]
        # 否则保留所有候选（至少有一个模糊匹配）

    for index in candidate_indices:
        line = cleaned_lines[index]

        year_column_value = extract_ocr_year_column_value(cleaned_lines, index, indicator)
        if year_column_value:
            return year_column_value

        if indicator["field_id"] == "G_Q_002":
            after = " ".join(cleaned_lines[index : min(len(cleaned_lines), index + 12)])
            person_matches = [
                (value, unit)
                for value, unit in ordered_number_unit_matches(after, filter_units=False)
                if unit in {"人", "名"}
            ]
            if person_matches:
                value, unit = person_matches[0]
                return value, unit, "ocr_independent_director_rule"
            raw_matches = ordered_number_unit_matches(after, filter_units=False)
            for value, unit in raw_matches:
                if not unit and float(value) <= 30:
                    return value, unit, "ocr_independent_director_rule"

        if indicator["field_id"] == "G_Q_003":
            around = " ".join(cleaned_lines[max(0, index - 4) : min(len(cleaned_lines), index + 5)])
            percentage_matches = [
                (value, unit)
                for value, unit in ordered_number_unit_matches(around, filter_units=False)
                if unit in {"%", "％"}
            ]
            if percentage_matches:
                value, unit = percentage_matches[0]
                return value, unit, "ocr_nearby_percentage_rule"

        same_line_matches = ordered_number_unit_matches(line, indicator)
        if same_line_matches:
            value, unit = same_line_matches[0]
            return value, unit, "ocr_same_line_rule"

        if indicator["field_id"] in {"E_Q_001", "E_Q_002", "E_Q_003"} and (
            "温室气体排放" in line or "范围一" in line or "范围二" in line
        ):
            local_window = " ".join(cleaned_lines[index : min(len(cleaned_lines), index + 12)])
            scope_matches = ordered_number_unit_matches(local_window, indicator)
            if indicator["field_id"] == "E_Q_001" and "范围一" in local_window and "范围二" in local_window and len(scope_matches) >= 2:
                total = sum(float(value.replace(",", "")) for value, _unit in scope_matches[:2])
                return format_numeric_value(total), scope_matches[0][1], "ocr_scope_sum_rule"
            if indicator["field_id"] == "E_Q_002" and scope_matches:
                return scope_matches[0][0], scope_matches[0][1], "ocr_scope_line_rule"
            if indicator["field_id"] == "E_Q_003" and len(scope_matches) >= 2:
                return scope_matches[1][0], scope_matches[1][1], "ocr_scope_line_rule"

        start = index
        while start > 0 and is_ocr_metric_label_line(cleaned_lines[start - 1]):
            start -= 1
        end = index
        while end + 1 < len(cleaned_lines) and is_ocr_metric_label_line(cleaned_lines[end + 1]):
            end += 1

        label_position = index - start
        metric_name = indicator.get("metric_name_cn", "")
        if ("危险" in metric_name or indicator.get("field_id") == "E_Q_013") and "危险" in line and line.find("危险") > 0:
            label_position += 1

        following: list[str] = []
        required_match_count = max(label_position + 1, end - start + 1)
        for cursor in range(end + 1, min(len(cleaned_lines), end + 10)):
            current = cleaned_lines[cursor]
            matches_before = ordered_number_unit_matches(" ".join(following), filter_units=False)
            if (
                following
                and is_ocr_metric_label_line(current)
                and matches_before
                and len(matches_before) >= required_match_count
            ):
                break
            following.append(current)
            matches_so_far = ordered_number_unit_matches(" ".join(following), filter_units=False)
            if len(matches_so_far) >= required_match_count:
                break

        value_matches = ordered_number_unit_matches(" ".join(following), filter_units=False)
        if len(value_matches) > label_position:
            value, unit = value_matches[label_position]
            # ── v1.3: 单位感知回退 ——
            # 如果位置配对的值单位不兼容，在后续值中搜索兼容单位的值
            if unit and not unit_is_compatible(unit, indicator):
                # 在 label_position ± 3 范围内搜索兼容单位的值
                search_start = max(0, label_position - 2)
                search_end = min(len(value_matches), label_position + 4)
                found_alt = False
                for alt_idx in range(search_start, search_end):
                    alt_value, alt_unit = value_matches[alt_idx]
                    if alt_unit and unit_is_compatible(alt_unit, indicator):
                        value, unit = alt_value, alt_unit
                        found_alt = True
                        break
                # 如果仍然没找到兼容单位，扩大搜索到所有后续值
                if not found_alt:
                    for alt_idx, (alt_value, alt_unit) in enumerate(value_matches):
                        if alt_unit and unit_is_compatible(alt_unit, indicator):
                            value, unit = alt_value, alt_unit
                            found_alt = True
                            break
                if not found_alt:
                    return None  # 无法找到兼容单位的值
            if not unit:
                local_window = " ".join(cleaned_lines[index : min(len(cleaned_lines), index + 9)])
                local_matches = ordered_number_unit_matches(local_window, indicator)
                if local_matches:
                    local_value, local_unit = local_matches[0]
                    return local_value, local_unit, "ocr_local_window_rule"
            if not unit or unit_is_compatible(unit, indicator):
                return value, unit, "ocr_line_pair_rule_v1.3"

        local_window = " ".join(cleaned_lines[index : min(len(cleaned_lines), index + 4)])
        local_matches = ordered_number_unit_matches(local_window, indicator)
        if local_matches:
            value, unit = local_matches[0]
            return value, unit, "ocr_local_window_rule"
    return None


def ocr_short_label_hits(page: dict[str, Any], indicator: dict[str, Any]) -> list[str]:
    if page.get("text_source") != "ocr_text" or indicator.get("metric_type") != "quantitative":
        return []
    text = page.get("text", "")
    if not text or not any(cue in text for cue in DATA_PAGE_CUES + STRUCTURED_VALUE_PAGE_CUES):
        return []
    lines = [normalize_text(line) for line in page.get("ocr_lines") or [] if normalize_text(line)]
    short_label_map = {
        "E_Q_007": ["电力"],
        "E_Q_009": ["水"],
    }
    hits = []
    for label in short_label_map.get(indicator["field_id"], []):
        if any(line == label for line in lines):
            hits.append(label)
    return hits


def cells_for_table_row(row: dict[str, Any]) -> list[str]:
    cells = row.get("cells") or []
    if not cells:
        cells = [part.strip() for part in row.get("row_text", "").split("|")]
    return [normalize_text(str(cell or "")) for cell in cells if normalize_text(str(cell or ""))]


def infer_unit_from_cells(cells: list[str], cell_index: int, indicator: dict[str, Any]) -> str:
    accepted_units = accepted_units_for(indicator)
    window = cells[max(0, cell_index - 2) : min(len(cells), cell_index + 2)]
    for cell in window:
        for unit in accepted_units:
            if unit and unit in cell:
                return unit
    joined = " ".join(window)
    matches = list(re.finditer(UNIT_PATTERN, joined, re.I))
    return matches[-1].group(0) if matches else ""


def _compute_row_context_score(row: dict[str, Any], indicator: dict[str, Any]) -> float:
    """v2.3: 计算指标别名在表格行中的上下文特异性得分。

    得分反映该行文本中指标别名的匹配质量：
    - 主名称精确命中 → 80
    - 特定别名（≥4字）命中 → 50
    - 模糊匹配（≥92%相似度，≥4字）→ 20
    - 无任何别名命中 → 0

    用于 extract_number_from_table_row 的强制门控：
    得分 < 20 时拒绝提取（防止无标签行被单位兼容性带偏）。
    """
    row_text = row.get("row_text", "")
    cell_texts = " ".join(cells_for_table_row(row))
    combined = f"{row_text} {cell_texts}"

    primary_name = indicator.get("metric_name_cn", "")
    indicator_aliases = [primary_name] + split_aliases(
        indicator.get("aliases_cn", ""), indicator.get("aliases_en", "")
    )
    extra_terms = EXTRA_TERMS.get(indicator["field_id"], [])

    # 1. 主名称精确命中（最高优先级）
    if primary_name and len(primary_name) >= 4 and primary_name in row_text:
        return 80.0
    if primary_name and len(primary_name) >= 4 and primary_name in cell_texts:
        return 70.0

    # 2. 别名精确命中
    for alias in sorted(indicator_aliases, key=len, reverse=True):
        if len(alias) >= 4 and alias in row_text:
            return 50.0
        if len(alias) >= 4 and alias in cell_texts:
            return 45.0

    # 3. EXTRA_TERMS 命中
    for term in extra_terms:
        if len(term) >= 3 and term in row_text:
            return 35.0
        if len(term) >= 3 and term in cell_texts:
            return 30.0

    # 4. 模糊匹配
    for alias in sorted(indicator_aliases, key=len, reverse=True):
        if len(alias) >= 4 and fuzz.partial_ratio(alias, combined) >= 92:
            return 20.0
    for term in extra_terms:
        if len(term) >= 4 and fuzz.partial_ratio(term, combined) >= 92:
            return 15.0

    return 0.0


def extract_number_from_table_row(row: dict[str, Any], indicator: dict[str, Any]) -> tuple[str, str, str] | None:
    """v2.3 增强版表格行数值抽取：强制上下文门控 + 年列感知 + 单位加权强化"""
    if indicator["metric_type"] != "quantitative":
        return None
    row_text = row.get("row_text", "")
    header_context = row.get("header_context", "")
    zero_value = explicit_zero_value(f"{row_text} {header_context}", indicator)
    if zero_value:
        return zero_value[0], zero_value[1], "table_explicit_zero"

    cells = cells_for_table_row(row)
    if not cells:
        return None

    # ── v2.3: 强制上下文门控 ──
    # 计算行级上下文特异性得分，低于阈值则拒绝提取
    context_score = _compute_row_context_score(row, indicator)
    if context_score < 20.0:
        # 行中无任何指标别名命中 → 不提取数值（防止无标签行被单位兼容性带偏）
        return None

    # ── v1.3: 检测年列位置 ──
    year_columns: set[int] = set()
    for cell_index, cell in enumerate(cells):
        if re.search(r"\b20(?:23|24|25|26)\b", cell):
            year_columns.add(cell_index)
    # 如果 header_context 中有年份标记，也尝试从 headers 推断
    if header_context:
        header_parts = header_context.split("||")
        for hp_idx, hp in enumerate(header_parts):
            if re.search(r"20(?:23|24|25|26)", hp):
                # 将 header 中的位置映射到 cells（近似）
                for ci in range(max(0, hp_idx - 1), min(len(cells), hp_idx + 4)):
                    year_columns.add(ci)

    # ── v2.3: 行标签-指标匹配质量（使用已计算的 context_score）──
    # context_score 已在函数入口处计算，直接使用
    row_label_match_bonus = context_score

    # ── v1.3: 增强评分 ──
    accepted_units = accepted_units_for(indicator)
    scored: list[tuple[float, str, str, int]] = []
    for cell_index, cell in enumerate(cells):
        for match in non_year_number_matches(cell):
            value = (match.group("value") or "").replace(",", "")
            unit = match.group("unit") or infer_unit_from_cells(cells, cell_index, indicator)
            numeric_value = float(value or "0")

            # 基础分：后列（通常为数值列）优先
            score = 8.0 + cell_index * 1.5

            # ── v1.3: 年列大幅加分 ──
            if cell_index in year_columns:
                score += 20.0
            # 如果有年列但当前不在年列，略降分
            elif year_columns and cell_index < max(year_columns):
                score -= 5.0  # 年列之前可能是旧年份

            # ── v1.3: 单位匹配强加权 ──
            if unit and unit_is_compatible(unit, indicator):
                score += 55.0  # 从 45 提升到 55
                # 精确单位匹配再加分
                if unit in accepted_units:
                    score += 12.0
            elif unit:
                score -= 45.0  # 从 -35 加强到 -45（不兼容单位的惩罚加重）
            elif indicator["value_type"] in ["integer", "number"]:
                score += 1.0  # 无单位时几乎不加分
            # 百分比/货币类型特殊处理
            if indicator["value_type"] == "percentage" and unit in ["%", "％"]:
                score += 35.0
            if indicator["value_type"] == "currency" and unit in ["元", "万元", "亿元", "人民币"]:
                score += 35.0
            # 过滤年份值
            if 1900 <= numeric_value <= 2100 and not unit:
                score -= 50.0

            # ── v1.3: 行标签匹配加分 ──
            score += row_label_match_bonus

            scored.append((score, value, unit, cell_index))

    if not scored:
        return None

    # 排序：得分最高的优先
    scored.sort(key=lambda item: item[0], reverse=True)

    # ── v2.4: 语义单位后验证 ──
    # 从最高分开始，跳过语义类型错误的单元格
    _time_units = {"小时", "学时", "分钟", "人时", "人次", "h", "hours", "hour"}
    _people_units = {"人", "名", "位", "person"}
    _mass_units = {"吨", "万吨", "t", "kg", "千克", "吨标煤", "吨标准煤"}
    norm_unit = (indicator.get("unit_normalized") or "").lower()
    # 判断指标期望的单位类型
    expects_time = any(tu in norm_unit for tu in _time_units)
    expects_people = any(pu in norm_unit for pu in _people_units) and not expects_time
    expects_pct = norm_unit == "%"
    best_idx = 0
    while best_idx < len(scored):
        _, val, unt, _ = scored[best_idx]
        skip = False
        if expects_time and unt in (_people_units - {"人时", "人次"}):
            skip = True  # 时间类指标不应有纯人员单位（允许复合单位如人时/人次）
        elif expects_people and unt in (_time_units | _mass_units):
            skip = True
        elif expects_pct and unt in (_people_units | _mass_units):
            skip = True
        if not skip:
            break
        best_idx += 1

    if best_idx >= len(scored):
        return None  # 所有候选都有语义错误的单位类型

    best_score, value, unit, best_cell_index = scored[best_idx]

    # ── v1.3: 增强过滤 ──
    if unit and not unit_is_compatible(unit, indicator):
        return None
    # 无单位且分数太低 → 拒绝
    if not unit and best_score < 12.0:
        return None
    # 有年列但选中了年列之前的、无单位的低分项 → 拒绝
    if year_columns and best_cell_index < max(year_columns) - 1 and not unit and best_score < 30.0:
        return None

    return value, unit, "table_cell_year_unit_v1.3"


VALUE_MISSING_CUES = [
    "具体金额未明确披露",
    "具体人数未明确披露",
    "具体时长未明确披露",
    "未明确披露",
    "未披露具体",
    "未披露",
    "未单独披露",
]


def value_status_for(indicator: dict[str, Any], value: str, unit: str, evidence_text: str, candidate_status: str) -> str:
    if candidate_status == "no_candidate":
        return "no_evidence"
    if indicator["metric_type"] != "quantitative":
        return "not_required"
    if value and unit_is_compatible(unit, indicator):
        return "exact_value_candidate"
    if any(cue in evidence_text for cue in VALUE_MISSING_CUES):
        return "value_missing_but_disclosed"
    if value:
        return "ambiguous_value_candidate"
    return "needs_value_review"


def confidence_from_score(score: float, evidence_type: str, has_value: bool, metric_type: str) -> float:
    base = min(score / 80, 0.88)
    if evidence_type == "native_table":
        base += 0.06
    if metric_type == "quantitative" and has_value:
        base += 0.05
    if metric_type == "qualitative":
        base += 0.04
    return round(max(0.05, min(base, 0.96)), 3)


def recommended_status(candidate_status: str, metric_type: str, value: str, source_text: str) -> str:
    if candidate_status == "no_candidate":
        return "not_found_review"
    if metric_type == "quantitative" and not value:
        return "needs_value_review"
    if len(source_text) < 20:
        return "needs_evidence_review"
    return "candidate_disclosed_review"


def candidate_disclosure_class(indicator: dict[str, Any], candidate_status: str) -> str:
    if candidate_status != "candidate_found":
        return "no_candidate"
    layer = indicator.get("indicator_layer") or "core"
    if layer == "alternative":
        return "alternative_disclosed"
    if layer == "component":
        return "component_disclosed"
    if layer == "industry_specific":
        return "industry_specific_disclosed"
    if layer == "qualitative_support":
        return "support_evidence_disclosed"
    return "primary_disclosed"


def unit_is_compatible(unit: str, indicator: dict[str, Any]) -> bool:
    if indicator["metric_type"] != "quantitative":
        return True
    if not unit:
        return False
    accepted_units = accepted_units_for(indicator)
    if indicator["field_id"] in {"E_Q_001", "E_Q_002", "E_Q_003"} and unit == "吨":
        return True
    if unit in accepted_units or unit == indicator.get("unit_normalized", ""):
        return True
    generic_units = {"吨", "万吨", "人", "名", "次", "场", "件", "%", "％"}
    if any(accepted and accepted not in generic_units and accepted in unit for accepted in accepted_units):
        return True
    if any(accepted and unit not in generic_units and unit in accepted for accepted in accepted_units):
        return True
    if indicator["value_type"] == "percentage" and unit in ["%", "％"]:
        return True
    if indicator["value_type"] == "currency" and unit in ["元", "万元", "亿元", "人民币"]:
        return True
    return False


def parse_numeric_value(value: Any) -> float | None:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def format_numeric_value(value: float) -> str:
    return f"{value:.6f}".rstrip("0").rstrip(".")


def standardize_value_unit(value: str, unit: str, indicator: dict[str, Any]) -> tuple[str, str]:
    if indicator["metric_type"] != "quantitative":
        return "", ""
    numeric = parse_numeric_value(value)
    if numeric is None:
        return "", ""
    raw_unit = unit or ""
    field_id = indicator["field_id"]
    unit_normalized = indicator.get("unit_normalized", "")

    if indicator["value_type"] == "percentage" or raw_unit in ["%", "％"]:
        return format_numeric_value(numeric), "%"

    if indicator["value_type"] == "currency" or field_id in {"E_Q_015", "S_Q_017"}:
        if "亿元" in raw_unit:
            return format_numeric_value(numeric * 10000), "万元"
        if "万元" in raw_unit:
            return format_numeric_value(numeric), "万元"
        if raw_unit in {"元", "人民币"}:
            return format_numeric_value(numeric / 10000), "万元"

    if field_id in {"E_Q_001", "E_Q_002", "E_Q_003"} or unit_normalized == "tCO2e":
        if "万吨" in raw_unit:
            return format_numeric_value(numeric * 10000), "tCO2e"
        if "千克" in raw_unit or "kgCO2e" in raw_unit:
            return format_numeric_value(numeric / 1000), "tCO2e"
        if any(token in raw_unit for token in ["吨", "tCO2e", "吨CO2e", "吨二氧化碳当量"]):
            return format_numeric_value(numeric), "tCO2e"

    if field_id == "E_Q_005":
        if "千克" in raw_unit or "kgCO2e" in raw_unit:
            return format_numeric_value(numeric / 1000), "tCO2e/unit"
        if "吨" in raw_unit or "tCO2e" in raw_unit:
            return format_numeric_value(numeric), "tCO2e/unit"

    if field_id == "E_Q_006":
        if "万吨" in raw_unit and ("标煤" in raw_unit or "标准煤" in raw_unit):
            return format_numeric_value(numeric * 10000), "吨标准煤"
        if "吨标煤" in raw_unit or "吨标准煤" in raw_unit:
            return format_numeric_value(numeric), "吨标准煤"

    if field_id == "E_Q_007":
        if "万" in raw_unit and ("千瓦时" in raw_unit or "kWh" in raw_unit):
            return format_numeric_value(numeric * 10000), "kWh"
        if "MWh" in raw_unit:
            return format_numeric_value(numeric * 1000), "kWh"
        if "千瓦时" in raw_unit or "kWh" in raw_unit:
            return format_numeric_value(numeric), "kWh"

    if field_id == "E_Q_009":
        if "万" in raw_unit and ("立方米" in raw_unit or "吨" in raw_unit or "m3" in raw_unit):
            return format_numeric_value(numeric * 10000), "m3"
        if "立方米" in raw_unit or raw_unit in {"m3", "吨"}:
            return format_numeric_value(numeric), "m3"

    if field_id in {"E_Q_012", "E_Q_013"}:
        if "万吨" in raw_unit:
            return format_numeric_value(numeric * 10000), "吨"
        if raw_unit in {"吨", "t"}:
            return format_numeric_value(numeric), "吨"
        if "kg" in raw_unit or "千克" in raw_unit:
            return format_numeric_value(numeric / 1000), "吨"

    if unit_normalized == "person" or raw_unit in {"人", "名"}:
        return format_numeric_value(numeric), "人"
    if raw_unit in {"人次", "次", "小时", "学时", "件", "场"}:
        return format_numeric_value(numeric), raw_unit
    return "", ""


def keep_candidate(record: dict[str, Any], indicator: dict[str, Any]) -> bool:
    if record["candidate_status"] != "candidate_found":
        return True
    if indicator["metric_type"] != "quantitative":
        return True
    if record.get("value_candidate"):
        if record.get("unit_raw_candidate") or record.get("unit_standardized_candidate"):
            return True
        text = f"{record.get('source_text', '')} {record.get('source_table_cell', '')}"
        return explicit_zero_value(text, indicator) is not None
    text = f"{record.get('source_text', '')} {record.get('source_table_cell', '')}"
    return explicit_zero_value(text, indicator) is not None


def value_is_known_false_positive(value: str, unit: str, indicator: dict[str, Any], evidence_text: str) -> bool:
    if indicator["field_id"] == "E_Q_009":
        if re.search(rf"{re.escape(value)}\s*{re.escape(unit or '')}\s*(竹浆|技改|项目|产能)", evidence_text):
            return True
        if "20万吨竹浆" in evidence_text and value == "20":
            return True
    return False


def build_candidate_record(
    sample: dict[str, Any],
    indicator: dict[str, Any],
    rank: int,
    status: str,
    evidence_type: str,
    score: float,
    source_page: str,
    source_physical_page: str,
    source_report_page_candidates: str,
    source_text: str,
    source_table_cell: str,
    match_terms: list[str],
    table_row: dict[str, Any] | None = None,
    page_payload: dict[str, Any] | None = None,
) -> dict[str, Any]:
    value_method = "text_rule"
    if evidence_type == "native_table" and table_row:
        table_value = extract_number_from_table_row(table_row, indicator)
        if table_value:
            value, unit, value_method = table_value
        else:
            fallback_text = table_row.get("row_text", "") if table_row else source_table_cell or source_text
            value, unit = extract_number(fallback_text, indicator)
            value_method = "table_fallback_text_rule" if value else "none"
    else:
        ocr_line_value = None
        if evidence_type == "ocr_text" and page_payload:
            ocr_line_value = extract_number_from_ocr_lines(list(page_payload.get("ocr_lines") or []), indicator)
        if ocr_line_value:
            value, unit, value_method = ocr_line_value
        else:
            value, unit = extract_number(source_text if evidence_type == "native_text" else source_table_cell or source_text, indicator)
            value_method = "text_rule" if value else "none"
    evidence_text = f"{source_text} {source_table_cell}"
    if value and value_is_known_false_positive(value, unit, indicator, evidence_text):
        value = ""
        unit = ""
        value_method = f"{value_method}_rejected_noise"
    value_standardized, unit_standardized = standardize_value_unit(value, unit, indicator)
    confidence = confidence_from_score(score, evidence_type, bool(value), indicator["metric_type"])
    next_status = recommended_status(status, indicator["metric_type"], value, source_text or source_table_cell)
    value_status = value_status_for(indicator, value, unit, evidence_text, status)
    needs_review = "yes"
    if status == "no_candidate":
        reason = "未找到足够匹配的候选证据，需要人工确认not_found或扩展别名。"
    elif indicator["metric_type"] == "quantitative" and not value:
        reason = "找到相关文本但未识别到数值，需要LLM/人工判断是否为定性描述或表格解析漏值。"
    elif confidence < 0.55:
        reason = "规则置信度偏低，需要LLM校验证据相关性。"
    else:
        reason = "候选证据可用，仍需LLM或人工确认字段口径、年份和单位。"
    return {
        "sample_id": sample["sample_id"],
        "stock_code": sample["stock_code"],
        "short_name": sample["short_name"],
        "report_type": sample["report_type"],
        "field_id": indicator["field_id"],
        "dimension": indicator["dimension"],
        "metric_name_cn": indicator["metric_name_cn"],
        "metric_type": indicator["metric_type"],
        "value_type": indicator["value_type"],
        "indicator_layer": indicator.get("indicator_layer", "core"),
        "primary_indicator_id": indicator.get("primary_indicator_id", ""),
        "rating_role": indicator.get("rating_role", ""),
        "alternative_status_policy": indicator.get("alternative_status_policy", ""),
        "scoring_denominator_policy": indicator.get("scoring_denominator_policy", ""),
        "candidate_status": status,
        "candidate_disclosure_class": candidate_disclosure_class(indicator, status),
        "candidate_rank": rank,
        "evidence_type_candidate": evidence_type,
        "value_candidate": value,
        "unit_raw_candidate": unit,
        "value_standardized_candidate": value_standardized,
        "unit_standardized_candidate": unit_standardized,
        "value_status": value_status,
        "value_extraction_method": value_method,
        "source_page": source_page,
        "source_physical_page": source_physical_page,
        "source_report_page_candidates": source_report_page_candidates,
        "source_text": source_text,
        "source_table_cell": source_table_cell,
        "match_terms": ";".join(match_terms[:8]),
        "rule_score": round(score, 2),
        "confidence_rule": confidence,
        "needs_llm_review": needs_review,
        "review_reason": reason,
        "recommended_next_status": next_status,
        "extractor_version": EXTRACTOR_VERSION,
        "pdf_path": sample["pdf_path"],
    }


def deduplicate_cross_indicator(
    records: list[dict[str, Any]],
    indicator_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    """v2.3: 跨指标去重安全网。

    当多个指标从同一页提取相同数值时，按别名匹配质量排他性分配。
    质量显著落后的候选（差距≥20分）→ 标记为 no_candidate。
    """
    # 按 (sample_id, source_physical_page, value_candidate, source_fingerprint) 分组
    # v2.3.1: 加入源文本指纹，防止不同行/段落的同值被误杀
    def _source_fingerprint(rec: dict[str, Any]) -> str:
        """取源文本前80字符作为指纹，区分不同表格行/段落。"""
        src = (rec.get("source_table_cell") or rec.get("source_text") or "")[:80]
        return src.strip()

    conflict_groups: dict[tuple[str, str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for rec in records:
        if rec.get("candidate_status") != "candidate_found":
            continue
        if rec.get("metric_type") != "quantitative":
            continue
        value = rec.get("value_candidate", "")
        if not value:
            continue
        page = rec.get("source_physical_page", "")
        # 仅当两个候选来自同一页、同一值、且同一源文本时才去重
        key = (rec["sample_id"], page, value, _source_fingerprint(rec))
        conflict_groups[key].append(rec)

    invalidated_count = 0
    for key, group in conflict_groups.items():
        if len(group) <= 1:
            continue

        # 为每个候选计算特异性得分
        scored_group: list[tuple[float, dict[str, Any]]] = []
        for rec in group:
            indicator = indicator_map.get(rec["field_id"], {})
            source_text = (
                (rec.get("source_text") or "")
                + " "
                + (rec.get("source_table_cell") or "")
            )

            specificity = 0.0

            # 1. 主名称命中
            primary = indicator.get("metric_name_cn", "")
            if primary and len(primary) >= 4 and primary in source_text:
                specificity += 50.0

            # 2. 别名命中
            aliases = split_aliases(
                indicator.get("aliases_cn", ""),
                indicator.get("aliases_en", ""),
            )
            alias_hits = sum(
                1 for a in aliases if a and len(a) >= 3 and a in source_text
            )
            specificity += alias_hits * 15.0

            # 3. 单位兼容性
            unit = rec.get("unit_raw_candidate", "")
            if unit and unit_is_compatible(unit, indicator):
                specificity += 20.0

            # 4. 现有规则分数
            specificity += float(rec.get("rule_score", 0)) * 0.5

            scored_group.append((specificity, rec))

        scored_group.sort(key=lambda x: x[0], reverse=True)
        best_score, best_rec = scored_group[0]

        # 质量显著落后的候选 → 无效化（v2.3.1: 阈值从20提高到40）
        for score, rec in scored_group[1:]:
            if best_score - score >= 40.0:
                rec["candidate_status"] = "no_candidate"
                rec["candidate_disclosure_class"] = "no_candidate"
                rec["value_candidate"] = ""
                rec["unit_raw_candidate"] = ""
                rec["value_standardized_candidate"] = ""
                rec["unit_standardized_candidate"] = ""
                rec["value_status"] = "needs_value_review"
                rec["value_extraction_method"] = (
                    f"{rec.get('value_extraction_method', '')}_cross_dedup"
                )
                rec["confidence_rule"] = "0.05"
                rec["review_reason"] = (
                    f"v2.3 cross-dedup: same value ({key[2]}) on page {key[1]} "
                    f"claimed by {best_rec['field_id']} (score={best_score:.1f} vs {score:.1f})"
                )
                invalidated_count += 1

    if invalidated_count:
        print(f"  Cross-indicator dedup: invalidated {invalidated_count} ambiguous candidates")

    return records


def find_candidates_for_indicator(
    sample: dict[str, Any],
    pdf_payload: dict[str, Any],
    indicator: dict[str, Any],
) -> list[dict[str, Any]]:
    terms = terms_for_indicator(indicator)
    primary_name = indicator.get("metric_name_cn", "")
    primary_aliases = split_aliases(
        indicator.get("aliases_cn", ""), indicator.get("aliases_en", "")
    )

    table_candidates: list[tuple[float, dict[str, Any], list[str]]] = []
    for row in pdf_payload["table_rows"]:
        row_text = row.get("row_text", "")
        scored_text = f"{row.get('header_context', '')} {row_text}"
        base_score, hits = page_score(row_text, terms)
        # ── v1.3: 行标签语义锚定 ──
        # 主指标名称精确匹配 → 大幅加分，抑制邻近行模糊匹配
        row_identity_bonus = 0.0
        if primary_name and len(primary_name) >= 4 and primary_name in row_text:
            row_identity_bonus = 35.0
        elif any(alias and len(alias) >= 4 and alias in row_text for alias in primary_aliases[:5]):
            row_identity_bonus = 20.0
        score = (
            base_score
            + row_identity_bonus
            + evidence_page_adjustment(scored_text, indicator)
            + evidence_structure_bonus(scored_text, indicator)
        )
        if base_score > 0 and score > 0:
            table_candidates.append((score + 8, row, hits))

    text_candidates: list[tuple[float, dict[str, Any], list[str]]] = []
    for page in pdf_payload["pages"]:
        base_score, hits = page_score(page["text"], terms)
        short_label_hits = []
        if base_score <= 0:
            short_label_hits = ocr_short_label_hits(page, indicator)
            if short_label_hits:
                base_score = 18 + len(short_label_hits) * 4
                hits = short_label_hits
        score = (
            base_score
            + evidence_page_adjustment(page["text"], indicator)
            + evidence_structure_bonus(page["text"], indicator)
        )
        # ── v2.4: G_Q_001 董事会总人数模式偏好 ──
        # "董事会由X名董事组成" 表示总数，"董事会人数" 在附录中可能是子项
        if indicator.get("field_id") == "G_Q_001":
            page_text = page.get("text", "")
            if re.search(r"董事会由.{0,15}名董事组成", page_text):
                score += 80.0  # 强偏好完整的董事会描述模式（总数）
            elif "董事会由" in page_text:
                score += 40.0
            # 附录表中 "董事会人数" 可能只是子计数（如执行董事人数）
            if re.search(r"董事会指标.{0,30}董事会人数", page_text):
                score -= 50.0  # 附录子表模式，几乎肯定不是总数
        if base_score > 0 and score > 0:
            text_candidates.append((score, page, hits))

    records: list[dict[str, Any]] = []
    rank = 1
    for score, row, hits in sorted(table_candidates, key=lambda item: item[0], reverse=True)[:2]:
        source = row["row_text"]
        source_pages, physical_page, report_page_candidates = source_page_reference(pdf_payload, row["page"])
        table_context = (
            f"header: {row.get('header_context', '')} || "
            f"table_{row['table_index']}_row_{row['row_index']}: {source}"
        )
        record = build_candidate_record(
            sample,
            indicator,
            rank,
            "candidate_found",
            "native_table",
            score,
            source_pages,
            physical_page,
            report_page_candidates,
            table_context,
            table_context,
            hits,
            row,
        )
        if keep_candidate(record, indicator):
            records.append(record)
            rank += 1

    text_limit = 2 if records else 3
    for score, page, hits in sorted(text_candidates, key=lambda item: item[0], reverse=True)[:text_limit]:
        snippet = snippet_around(page["text"], [h.split("~")[0] for h in hits])
        physical_page = str(page["page"])
        report_page_candidates = ";".join(str(item) for item in page.get("report_page_candidates") or [])
        source_pages = ";".join(unique_preserve_order(([physical_page] if physical_page else []) + list(page.get("report_page_candidates") or [])))
        if any(existing["source_physical_page"] == physical_page and existing["source_text"] == snippet for existing in records):
            continue
        record = build_candidate_record(
            sample,
            indicator,
            rank,
            "candidate_found",
            page.get("text_source") or "native_text",
            score,
            source_pages,
            physical_page,
            report_page_candidates,
            snippet,
            "",
            hits,
            page_payload=page,
        )
        if keep_candidate(record, indicator):
            records.append(record)
            rank += 1

    if not records:
        records.append(
            build_candidate_record(
                sample,
                indicator,
                1,
                "no_candidate",
                "",
                0.0,
                "",
                "",
                "",
                "",
                "",
                [],
            )
        )
    return records[:3]


def write_csv(records: list[dict[str, Any]]) -> None:
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CANDIDATE_COLUMNS)
        writer.writeheader()
        writer.writerows(records)


def make_summaries(records: list[dict[str, Any]], samples: list[dict[str, Any]], indicators: list[dict[str, Any]]) -> dict[str, Any]:
    sample_summary: list[dict[str, Any]] = []
    indicator_summary: list[dict[str, Any]] = []
    by_sample = defaultdict(list)
    by_indicator = defaultdict(list)
    for record in records:
        by_sample[record["sample_id"]].append(record)
        by_indicator[record["field_id"]].append(record)

    for sample in samples:
        rows = by_sample[sample["sample_id"]]
        field_ids_with_candidate = {
            row["field_id"] for row in rows if row["candidate_status"] == "candidate_found"
        }
        candidate_rows = [r for r in rows if r["candidate_status"] == "candidate_found"]
        sample_summary.append(
            {
                "sample_id": sample["sample_id"],
                "stock_code": sample["stock_code"],
                "short_name": sample["short_name"],
                "report_type": sample["report_type"],
                "page_count": sample["page_count"],
                "candidate_field_count": len(field_ids_with_candidate),
                "p0_field_count": len(indicators),
                "candidate_field_rate": round(len(field_ids_with_candidate) / len(indicators), 4),
                "candidate_row_count": len(candidate_rows),
                "primary_disclosed_fields": len({r["field_id"] for r in candidate_rows if r["candidate_disclosure_class"] == "primary_disclosed"}),
                "alternative_disclosed_fields": len({r["field_id"] for r in candidate_rows if r["candidate_disclosure_class"] == "alternative_disclosed"}),
                "component_disclosed_fields": len({r["field_id"] for r in candidate_rows if r["candidate_disclosure_class"] == "component_disclosed"}),
                "industry_specific_disclosed_fields": len({r["field_id"] for r in candidate_rows if r["candidate_disclosure_class"] == "industry_specific_disclosed"}),
                "support_evidence_disclosed_fields": len({r["field_id"] for r in candidate_rows if r["candidate_disclosure_class"] == "support_evidence_disclosed"}),
                "table_candidate_rows": len([r for r in rows if r["evidence_type_candidate"] == "native_table"]),
                "text_candidate_rows": len([r for r in rows if r["evidence_type_candidate"] in {"native_text", "ocr_text"}]),
                "ocr_text_candidate_rows": len([r for r in rows if r["evidence_type_candidate"] == "ocr_text"]),
                "no_candidate_fields": len(indicators) - len(field_ids_with_candidate),
            }
        )

    indicator_map = {item["field_id"]: item for item in indicators}
    for field_id, rows in sorted(by_indicator.items()):
        found_samples = {row["sample_id"] for row in rows if row["candidate_status"] == "candidate_found"}
        indicator_summary.append(
            {
                "field_id": field_id,
                "dimension": indicator_map[field_id]["dimension"],
                "metric_name_cn": indicator_map[field_id]["metric_name_cn"],
                "metric_type": indicator_map[field_id]["metric_type"],
                "indicator_layer": indicator_map[field_id].get("indicator_layer", ""),
                "primary_indicator_id": indicator_map[field_id].get("primary_indicator_id", ""),
                "rating_role": indicator_map[field_id].get("rating_role", ""),
                "samples_with_candidate": len(found_samples),
                "sample_count": len(samples),
                "candidate_sample_rate": round(len(found_samples) / len(samples), 4),
                "table_rows": len([r for r in rows if r["evidence_type_candidate"] == "native_table"]),
                "text_rows": len([r for r in rows if r["evidence_type_candidate"] in {"native_text", "ocr_text"}]),
                "ocr_text_rows": len([r for r in rows if r["evidence_type_candidate"] == "ocr_text"]),
            }
        )

    return {
        "sample_summary": sample_summary,
        "indicator_summary": indicator_summary,
        "overall": {
            "sample_count": len(samples),
            "p0_indicator_count": len(indicators),
            "candidate_records": len(records),
            "candidate_found_rows": len([r for r in records if r["candidate_status"] == "candidate_found"]),
            "no_candidate_rows": len([r for r in records if r["candidate_status"] == "no_candidate"]),
            "fields_with_candidate_total": sum(item["candidate_field_count"] for item in sample_summary),
            "field_sample_pairs_total": len(samples) * len(indicators),
            "field_sample_candidate_rate": round(
                sum(item["candidate_field_count"] for item in sample_summary) / (len(samples) * len(indicators)),
                4,
            ),
            "evidence_type_counts": dict(Counter(r["evidence_type_candidate"] or "none" for r in records)),
            "candidate_disclosure_class_counts": dict(Counter(r["candidate_disclosure_class"] for r in records)),
            "indicator_layer_counts": dict(Counter(i.get("indicator_layer", "core") for i in indicators)),
        },
    }


def write_json(records: list[dict[str, Any]], summaries: dict[str, Any], samples: list[dict[str, Any]], indicators: list[dict[str, Any]]) -> None:
    payload = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "extractor_version": EXTRACTOR_VERSION,
        "pilot_sample_ids": [sample["sample_id"] for sample in samples],
        "samples": samples,
        "p0_indicators": indicators,
        "summaries": summaries,
        "records": records,
    }
    JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = min(max(len(str(cell.value or "")) for cell in col) + 2, 48)
        ws.column_dimensions[col[0].column_letter].width = max(10, max_len)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def xlsx_safe_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def append_rows(ws, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([xlsx_safe_value(row.get(col, "")) for col in columns])
    style_sheet(ws)


def write_xlsx(records: list[dict[str, Any]], summaries: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "候选抽取明细"
    append_rows(ws, records, CANDIDATE_COLUMNS)

    sample_ws = wb.create_sheet("样本汇总")
    sample_cols = list(summaries["sample_summary"][0].keys()) if summaries["sample_summary"] else []
    append_rows(sample_ws, summaries["sample_summary"], sample_cols)

    indicator_ws = wb.create_sheet("指标汇总")
    indicator_cols = list(summaries["indicator_summary"][0].keys()) if summaries["indicator_summary"] else []
    append_rows(indicator_ws, summaries["indicator_summary"], indicator_cols)

    overall_ws = wb.create_sheet("总体统计")
    overall_ws.append(["指标", "值"])
    for key, value in summaries["overall"].items():
        overall_ws.append([key, json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
    style_sheet(overall_ws)
    overall_ws.column_dimensions["A"].width = 30
    overall_ws.column_dimensions["B"].width = 90
    wb.save(XLSX_PATH)


def llm_prompt_for(record: dict[str, Any], indicator: dict[str, Any]) -> dict[str, Any]:
    schema = {
        "status": "disclosed|not_disclosed|not_found|not_applicable",
        "value": "数值或定性原文；不确定则为空",
        "unit_raw": "报告原始单位；没有则为空",
        "source_page": "证据页码",
        "source_text": "可验证原文片段",
        "evidence_type": "native_text|native_table|ocr_text|ocr_table|chart|manual",
        "applicability_basis": "not_applicable时填写行业、业务或报告边界依据；否则为空",
        "provenance_type": "reported|extracted|manual|estimated；抽取层默认不使用estimated",
        "gap_policy": "scoring_excluded|coverage_penalty|transparency_penalty|technical_review|use_value",
        "confidence": "0-1",
        "reason": "一句话说明判断依据",
    }
    user = (
        "请基于候选证据判断该ESG指标是否已披露，并按JSON输出。"
        "只能使用给定证据，不得猜测。若证据不足，输出not_found；若候选证据能证明该指标对该公司不适用，输出not_applicable。\n\n"
        f"公司：{record['short_name']}（{record['stock_code']}）\n"
        f"指标ID：{record['field_id']}\n"
        f"指标名称：{record['metric_name_cn']}\n"
        f"指标类型：{record['metric_type']}\n"
        f"定义：{indicator.get('definition','')}\n"
        f"标准单位：{indicator.get('unit_normalized','')}\n"
        f"别名：{indicator.get('aliases_cn','')}\n"
        f"页码：{record['source_page']}\n"
        f"候选证据类型：{record['evidence_type_candidate']}\n"
        f"候选数值：{record['value_candidate']} {record['unit_raw_candidate']}\n"
        f"候选原文/表格：{record['source_table_cell'] or record['source_text']}\n\n"
        f"输出JSON字段：{json.dumps(schema, ensure_ascii=False)}"
    )
    return {
        "custom_id": f"{record['sample_id']}_{record['field_id']}_{record['candidate_rank']}",
        "sample_id": record["sample_id"],
        "field_id": record["field_id"],
        "messages": [
            {
                "role": "system",
                "content": "你是ESG报告指标抽取审核员，强调证据可追溯、单位准确和不臆测。",
            },
            {"role": "user", "content": user},
        ],
        "expected_json_schema": schema,
    }


def write_llm_queue(records: list[dict[str, Any]], indicators: list[dict[str, Any]]) -> None:
    indicator_map = {item["field_id"]: item for item in indicators}
    queue_records = [
        record
        for record in records
        if record["candidate_status"] == "candidate_found" and int(record["candidate_rank"]) == 1
    ]
    with LLM_JSONL_PATH.open("w", encoding="utf-8") as handle:
        for record in queue_records:
            handle.write(json.dumps(llm_prompt_for(record, indicator_map[record["field_id"]]), ensure_ascii=False) + "\n")


def write_review_queue(records: list[dict[str, Any]]) -> None:
    review_columns = [
        "sample_id",
        "stock_code",
        "short_name",
        "report_type",
        "field_id",
        "dimension",
        "metric_name_cn",
        "metric_type",
        "candidate_status",
        "evidence_type_candidate",
        "value_candidate",
        "unit_raw_candidate",
        "source_page",
        "source_text",
        "source_table_cell",
        "confidence_rule",
        "recommended_next_status",
        "manual_status",
        "manual_value",
        "manual_unit_raw",
        "manual_source_page",
        "applicability_basis",
        "materiality_level",
        "provenance_type",
        "gap_policy",
        "manual_note",
        "reviewer",
    ]
    first_rank = [record for record in records if int(record["candidate_rank"]) == 1]
    wb = Workbook()
    ws = wb.active
    ws.title = "P0人工复核清单"
    ws.append(review_columns)
    for record in first_rank:
        row = {col: record.get(col, "") for col in review_columns}
        row["manual_status"] = ""
        row["manual_value"] = ""
        row["manual_unit_raw"] = ""
        row["manual_source_page"] = ""
        row["applicability_basis"] = ""
        row["materiality_level"] = ""
        row["provenance_type"] = "reported"
        row["gap_policy"] = ""
        row["manual_note"] = ""
        row["reviewer"] = ""
        ws.append([xlsx_safe_value(row.get(col, "")) for col in review_columns])
    style_sheet(ws)

    guide = wb.create_sheet("复核说明")
    guide.append(["字段", "说明"])
    notes = [
        ("manual_status", "填写 disclosed / not_disclosed / not_found / not_applicable / unsure。"),
        ("manual_value", "确认后的数值或定性原文；不要直接相信候选值。"),
        ("manual_unit_raw", "报告原始单位，如万吨、万元、人、%。"),
        ("manual_source_page", "确认后的PDF页码。"),
        ("applicability_basis", "not_applicable时必须填写行业、业务或报告边界依据。"),
        ("provenance_type", "reported;extracted;manual;estimated。抽取层默认不使用estimated。"),
        ("gap_policy", "scoring_excluded;coverage_penalty;transparency_penalty;technical_review。"),
        ("manual_note", "记录年份口径、单位换算、候选错误、需讨论等情况。"),
    ]
    for note in notes:
        guide.append(list(note))
    style_sheet(guide)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 90
    wb.save(REVIEW_XLSX_PATH)


def md_table(rows: list[dict[str, Any]], columns: list[str]) -> str:
    lines = ["| " + " | ".join(columns) + " |", "|" + "|".join(["---"] * len(columns)) + "|"]
    for row in rows:
        lines.append("| " + " | ".join(str(row.get(col, "")) for col in columns) + " |")
    return "\n".join(lines)


def write_report(summaries: dict[str, Any], samples: list[dict[str, Any]], elapsed_sec: float) -> None:
    overall = summaries["overall"]
    lines = [
        "# P0指标候选抽取试跑报告",
        "",
        f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 试跑设置",
        "",
        "- 试跑范围：按环境变量选择报告，39个P0指标。",
        "- 技术路线：先用PDF原生文本和表格抽取生成候选证据；原生文本为空时读取OCR缓存，并将证据标记为`ocr_text`；再进入LLM/人工校验。",
        "- 当前结果是候选层结果，不代表最终准确率或F1。",
        f"- 运行耗时：{round(elapsed_sec, 2)} 秒。",
        "",
        "## 样本选择",
        "",
        "| 样本ID | 股票代码 | 简称 | 类型 | 页数 | 选择理由 |",
        "|---|---|---|---|---:|---|",
    ]
    for sample in samples:
        lines.append(
            f"| {sample['sample_id']} | {sample['stock_code']} | {sample['short_name']} | "
            f"{sample['report_type']} | {sample['page_count']} | {sample['sampling_reason']} |"
        )
    lines.extend(
        [
            "",
            "## 总体结果",
            "",
            f"- 字段-样本组合：{overall['field_sample_pairs_total']} 个。",
            f"- 找到候选证据的组合：{overall['fields_with_candidate_total']} 个。",
            f"- 候选覆盖率：{overall['field_sample_candidate_rate']:.2%}。",
            f"- 候选明细行：{overall['candidate_records']} 行，其中候选命中 {overall['candidate_found_rows']} 行，未命中占位 {overall['no_candidate_rows']} 行。",
            f"- 证据类型分布：{json.dumps(overall['evidence_type_counts'], ensure_ascii=False)}。",
            "",
            "## 样本维度汇总",
            "",
            md_table(
                summaries["sample_summary"],
                [
                    "sample_id",
                    "stock_code",
                    "short_name",
                    "candidate_field_count",
                    "p0_field_count",
                    "candidate_field_rate",
                    "table_candidate_rows",
                    "text_candidate_rows",
                    "ocr_text_candidate_rows",
                    "no_candidate_fields",
                ],
            ),
            "",
            "## 候选覆盖较低的P0指标",
            "",
        ]
    )
    weak = sorted(summaries["indicator_summary"], key=lambda row: (row["samples_with_candidate"], row["field_id"]))[:12]
    lines.append(
        md_table(
            weak,
            ["field_id", "dimension", "metric_name_cn", "metric_type", "samples_with_candidate", "sample_count"],
        )
    )
    lines.extend(
        [
            "",
            "## 初步观察",
            "",
            "- 表格路径已经能为环境绩效、员工、董事会等指标提供一批结构化候选，但仍需要处理跨页表、合并表头和单位继承。",
            "- v0.6新增OCR缓存兜底，适用于GL010、GL014、GL020等图片型PDF；OCR证据需优先进入LLM/人工校验，避免把识别噪声当作最终数值。",
            "- CSR/ENV短报告会自然产生较多`no_candidate`，这对后续`not_found/not_disclosed`判断很重要。",
            "- 定性指标通常能通过别名检索找到候选页，但候选片段是否真正回答指标，需要LLM或人工校验。",
            "- 当前未启用OCR，因此低文本质量PDF暂不纳入本轮试跑；下一轮应单独做OCR小实验。",
            "",
            "## 产物位置",
            "",
            f"- 候选明细CSV：`{CSV_PATH}`",
            f"- 候选明细Excel：`{XLSX_PATH}`",
            f"- 候选明细JSON：`{JSON_PATH}`",
            f"- LLM校验JSONL：`{LLM_JSONL_PATH}`",
            f"- 人工复核清单：`{REVIEW_XLSX_PATH}`",
            f"- 页级文本缓存：`{PAGE_TEXT_DIR}`",
            "",
            "## 下一步",
            "",
            "1. 用人工标注模板复核每份报告的P0候选，确认候选是否为真实披露。",
            "2. 将候选结果和人工金标准对齐，计算字段级precision、recall、F1、页码准确率和单位准确率。",
            "3. 对未命中但人工标注有值的字段补充别名、章节词和表格定位规则。",
            "4. 对表格候选中单位缺失的行加入表头/脚注单位继承逻辑。",
        ]
    )
    REPORT_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    PAGE_TEXT_DIR.mkdir(parents=True, exist_ok=True)
    samples, indicators = load_inputs()

    all_terms: list[str] = []
    for indicator in indicators:
        all_terms.extend(split_aliases(indicator["metric_name_cn"], indicator.get("aliases_cn", ""), indicator.get("aliases_en", "")))
    global_clues = sorted(set(TABLE_CUES + all_terms), key=len, reverse=True)

    start = time.time()
    pdf_payloads: dict[str, dict[str, Any]] = {}
    for sample in samples:
        print(f"extracting {sample['sample_id']} {sample['stock_code']} {sample['short_name']}")
        pdf_payloads[sample["sample_id"]] = extract_pdf(sample, global_clues)

    records: list[dict[str, Any]] = []
    for sample in samples:
        payload = pdf_payloads[sample["sample_id"]]
        for indicator in indicators:
            records.extend(find_candidates_for_indicator(sample, payload, indicator))

    # ── v2.3: 跨指标去重 ──
    indicator_map = {item["field_id"]: item for item in indicators}
    records = deduplicate_cross_indicator(records, indicator_map)

    summaries = make_summaries(records, samples, indicators)
    write_csv(records)
    write_json(records, summaries, samples, indicators)
    write_xlsx(records, summaries)
    write_llm_queue(records, indicators)
    write_review_queue(records)
    elapsed = time.time() - start
    write_report(summaries, samples, elapsed)
    print(json.dumps({"elapsed_sec": round(elapsed, 2), **summaries["overall"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
