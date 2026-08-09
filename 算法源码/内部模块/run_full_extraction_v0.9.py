# -*- coding: utf-8 -*-
"""
ESG 报告全量指标候选抽取器 v0.9

升级内容（相对 v0.8c）：
1. 支持 80 指标全量抽取（P0/P1/P2 分层），通过 PILOT_PRIORITY 环境变量控制
2. 集成指标索引页反向定位模块
3. 保留 P0/P1/P2 分层标记，下游评分时区分处理
4. 使用独立 v0.9 输出目录

用法：
  # 全量80指标
  python run_full_extraction_v0.9.py

  # 仅P0（保持向后兼容）
  PILOT_PRIORITY=P0 python run_full_extraction_v0.9.py

  # P0+P1
  PILOT_PRIORITY=P0,P1 python run_full_extraction_v0.9.py

  # 只跑特定样本
  PILOT_SAMPLE_IDS=GL020 python run_full_extraction_v0.9.py
"""

import sys
from pathlib import Path

# ── 将原 v0.8c 脚本路径加入 sys.path 以便导入核心函数 ──
_SCRIPT_DIR = Path(__file__).resolve().parent
if str(_SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPT_DIR))

# 直接导入原脚本的全部函数（通过 exec 避免 main() 被触发）
import importlib.util as _importlib_util
_V08C_PATH = _SCRIPT_DIR / "run_p0_pilot_extraction.py"
_V08C_SPEC = _importlib_util.spec_from_file_location("run_p0_pilot_extraction_v08c", _V08C_PATH)
_V08C_MODULE = _importlib_util.module_from_spec(_V08C_SPEC)

# 在执行模块前，设置环境变量阻止 main() 运行
import os as _os
_original_name = _os.environ.get("_V08C_IMPORT_MODE", "")
_os.environ["_V08C_IMPORT_MODE"] = "1"

# 但我们不能直接 exec_module 因为那会运行 main()
# 更好的方式：读取原脚本源码，去掉 main() 调用后 exec
_V08C_SOURCE = _V08C_PATH.read_text(encoding="utf-8")
# 找到 if __name__ == "__main__": 并注释掉
import re as _re
_V08C_SOURCE = _re.sub(
    r'if\s+__name__\s*==\s*["\']__main__["\']\s*:',
    'if False and __name__ == "__main__":',
    _V08C_SOURCE
)
exec(_V08C_SOURCE, _V08C_MODULE.__dict__)

# 现在从模块中取函数
normalize_text = _V08C_MODULE.__dict__["normalize_text"]
unique_preserve_order = _V08C_MODULE.__dict__["unique_preserve_order"]
valid_report_page_number = _V08C_MODULE.__dict__["valid_report_page_number"]
infer_report_page_candidates = _V08C_MODULE.__dict__["infer_report_page_candidates"]
source_page_reference = _V08C_MODULE.__dict__["source_page_reference"]
load_ocr_payload_for_page = _V08C_MODULE.__dict__["load_ocr_payload_for_page"]
infer_ocr_report_page_candidates = _V08C_MODULE.__dict__["infer_ocr_report_page_candidates"]
split_aliases = _V08C_MODULE.__dict__["split_aliases"]
split_units = _V08C_MODULE.__dict__["split_units"]
terms_for_indicator = _V08C_MODULE.__dict__["terms_for_indicator"]
choose_sample_ids = _V08C_MODULE.__dict__["choose_sample_ids"]
extract_pdf = _V08C_MODULE.__dict__["extract_pdf"]
page_score = _V08C_MODULE.__dict__["page_score"]
evidence_page_adjustment = _V08C_MODULE.__dict__["evidence_page_adjustment"]
evidence_structure_bonus = _V08C_MODULE.__dict__["evidence_structure_bonus"]
snippet_around = _V08C_MODULE.__dict__["snippet_around"]
accepted_units_for = _V08C_MODULE.__dict__["accepted_units_for"]
default_unit_for_zero = _V08C_MODULE.__dict__["default_unit_for_zero"]
explicit_zero_value = _V08C_MODULE.__dict__["explicit_zero_value"]
unit_is_compatible = _V08C_MODULE.__dict__["unit_is_compatible"]
extract_number = _V08C_MODULE.__dict__["extract_number"]
extract_number_from_ocr_lines = _V08C_MODULE.__dict__["extract_number_from_ocr_lines"]
ocr_short_label_hits = _V08C_MODULE.__dict__["ocr_short_label_hits"]
cells_for_table_row = _V08C_MODULE.__dict__["cells_for_table_row"]
extract_number_from_table_row = _V08C_MODULE.__dict__["extract_number_from_table_row"]
value_status_for = _V08C_MODULE.__dict__["value_status_for"]
confidence_from_score = _V08C_MODULE.__dict__["confidence_from_score"]
recommended_status = _V08C_MODULE.__dict__["recommended_status"]
candidate_disclosure_class = _V08C_MODULE.__dict__["candidate_disclosure_class"]
keep_candidate = _V08C_MODULE.__dict__["keep_candidate"]
value_is_known_false_positive = _V08C_MODULE.__dict__["value_is_known_false_positive"]
build_candidate_record = _V08C_MODULE.__dict__["build_candidate_record"]
find_candidates_for_indicator = _V08C_MODULE.__dict__["find_candidates_for_indicator"]
write_csv = _V08C_MODULE.__dict__["write_csv"]
make_summaries = _V08C_MODULE.__dict__["make_summaries"]
write_json = _V08C_MODULE.__dict__["write_json"]
style_sheet = _V08C_MODULE.__dict__["style_sheet"]
append_rows = _V08C_MODULE.__dict__["append_rows"]
write_xlsx = _V08C_MODULE.__dict__["write_xlsx"]
llm_prompt_for = _V08C_MODULE.__dict__["llm_prompt_for"]
write_llm_queue = _V08C_MODULE.__dict__["write_llm_queue"]
write_review_queue = _V08C_MODULE.__dict__["write_review_queue"]
md_table = _V08C_MODULE.__dict__["md_table"]
write_report = _V08C_MODULE.__dict__["write_report"]
parse_numeric_value = _V08C_MODULE.__dict__["parse_numeric_value"]
format_numeric_value = _V08C_MODULE.__dict__["format_numeric_value"]
standardize_value_unit = _V08C_MODULE.__dict__["standardize_value_unit"]
number_match_is_noise = _V08C_MODULE.__dict__["number_match_is_noise"]
non_year_number_matches = _V08C_MODULE.__dict__["non_year_number_matches"]
year_sequence = _V08C_MODULE.__dict__["year_sequence"]
year_aware_number = _V08C_MODULE.__dict__["year_aware_number"]
direct_labeled_number = _V08C_MODULE.__dict__["direct_labeled_number"]
is_ocr_metric_label_line = _V08C_MODULE.__dict__["is_ocr_metric_label_line"]
ordered_number_unit_matches = _V08C_MODULE.__dict__["ordered_number_unit_matches"]
extract_ocr_year_column_value = _V08C_MODULE.__dict__["extract_ocr_year_column_value"]
infer_unit_near = _V08C_MODULE.__dict__["infer_unit_near"]
infer_unit_from_cells = _V08C_MODULE.__dict__["infer_unit_from_cells"]
deduplicate_cross_indicator = _V08C_MODULE.__dict__["deduplicate_cross_indicator"]

# 导入索引页解析器
from index_page_resolver import (
    resolve_index_target_pages,
    map_index_to_indicators,
    parse_index_page_from_ocr_json,
    index_resolver_summary,
)

import json as _json
from datetime import datetime as _datetime
from collections import Counter as _Counter
from collections import defaultdict as _defaultdict
import os as _v9_os
import time as _v9_time
from openpyxl import Workbook as _v9_Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as _XLSX_ILLEGAL_CHARACTERS_RE

_BASE_DIR = Path(__file__).resolve().parents[2]
_SAMPLE_JSON = Path(_v9_os.environ.get(
    "SAMPLE_JSON_PATH",
    str(_BASE_DIR / "算法源码" / "示例清单" / "示例样本清单.json"),
))
_INDICATOR_JSON = Path(_v9_os.environ.get(
    "ESG_INDICATOR_JSON",
    str(_BASE_DIR / "算法源码" / "配置" / "ESG指标体系.json"),
))
_OUT_DIR = Path(_v9_os.environ.get("PILOT_OUT_DIR", str(_BASE_DIR / "运行产物" / "候选抽取")))
_PAGE_TEXT_DIR = _OUT_DIR / "extracted_page_text"
_OCR_CACHE_DIR = Path(_v9_os.environ.get("OCR_CACHE_DIR", str(_BASE_DIR / "运行缓存" / "OCR")))
_OCR_JSON_DIR = _OCR_CACHE_DIR / "ocr_page_json"
_OCR_TEXT_DIR = _OCR_CACHE_DIR / "ocr_pages"
_RUN_LABEL = _v9_os.environ.get("PILOT_RUN_LABEL", "20份v0.9全量80指标")
_EXTRACTION_PRIORITIES = _v9_os.environ.get("PILOT_PRIORITY", "all").strip()
_ENABLE_INDEX_RESOLVER = _v9_os.environ.get("ENABLE_INDEX_RESOLVER", "1").strip() in {"1", "true", "yes", "y"}

_CSV_PATH = _OUT_DIR / f"全量指标候选抽取结果_{_RUN_LABEL}.csv"
_JSON_PATH = _OUT_DIR / f"全量指标候选抽取结果_{_RUN_LABEL}.json"
_XLSX_PATH = _OUT_DIR / f"全量指标候选抽取结果_{_RUN_LABEL}.xlsx"
_LLM_JSONL_PATH = _OUT_DIR / f"LLM校验提示队列_{_RUN_LABEL}.jsonl"
_REVIEW_XLSX_PATH = _OUT_DIR / f"全量指标人工复核清单_{_RUN_LABEL}.xlsx"
_REPORT_PATH = _OUT_DIR / "全量指标候选抽取试跑报告_v0.9.md"
_INDEX_REPORT_PATH = _OUT_DIR / "索引页反向定位报告_v0.9.md"

_EXTRACTOR_VERSION = "p0_candidate_pilot_v0.9"


def _xlsx_safe_value(value):
    if isinstance(value, str):
        return _XLSX_ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def safe_append_rows(ws, rows: list[dict[str, any]], columns: list[str]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([_xlsx_safe_value(row.get(col, "")) for col in columns])


def write_review_queue_safe(records: list[dict[str, any]]) -> None:
    """Write manual review workbook with XLSX-illegal PDF control chars removed."""
    review_columns = [
        "sample_id",
        "stock_code",
        "short_name",
        "report_type",
        "field_id",
        "dimension",
        "metric_name_cn",
        "metric_type",
        "candidate_status",
        "evidence_type_candidate",
        "value_candidate",
        "unit_raw_candidate",
        "source_page",
        "source_text",
        "source_table_cell",
        "confidence_rule",
        "recommended_next_status",
        "manual_status",
        "manual_value",
        "manual_unit_raw",
        "manual_source_page",
        "applicability_basis",
        "materiality_level",
        "provenance_type",
        "gap_policy",
        "manual_note",
        "reviewer",
    ]
    first_rank = [record for record in records if int(record["candidate_rank"]) == 1]
    rows = []
    for record in first_rank:
        row = {col: record.get(col, "") for col in review_columns}
        row["manual_status"] = ""
        row["manual_value"] = ""
        row["manual_unit_raw"] = ""
        row["manual_source_page"] = ""
        row["applicability_basis"] = ""
        row["materiality_level"] = ""
        row["provenance_type"] = "reported"
        row["gap_policy"] = ""
        row["manual_note"] = ""
        row["reviewer"] = ""
        rows.append(row)

    wb = _v9_Workbook()
    ws = wb.active
    ws.title = "P0人工复核清单"
    safe_append_rows(ws, rows, review_columns)
    style_sheet(ws)

    guide = wb.create_sheet("复核说明")
    guide.append(["字段", "说明"])
    notes = [
        ("manual_status", "填写 disclosed / not_disclosed / not_found / not_applicable / unsure。"),
        ("manual_value", "确认后的数值或定性原文；不要直接相信候选值。"),
        ("manual_unit_raw", "报告原始单位，如万吨、万元、人、%。"),
        ("manual_source_page", "确认后的PDF页码。"),
        ("applicability_basis", "not_applicable时必须填写行业、业务或报告边界依据。"),
        ("provenance_type", "reported;extracted;manual;estimated。抽取层默认不使用estimated。"),
        ("gap_policy", "scoring_excluded;coverage_penalty;transparency_penalty;technical_review。"),
        ("manual_note", "记录年份口径、单位换算、候选错误、需讨论等情况。"),
    ]
    for note in notes:
        guide.append([_xlsx_safe_value(item) for item in note])
    style_sheet(guide)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 90
    wb.save(_REVIEW_XLSX_PATH)

# ── 重定向 v0.8c 模块的路径到 v0.9 目录 ──
_V08C_MODULE.__dict__["OUT_DIR"] = _OUT_DIR
_V08C_MODULE.__dict__["PAGE_TEXT_DIR"] = _PAGE_TEXT_DIR
_V08C_MODULE.__dict__["CSV_PATH"] = _CSV_PATH
_V08C_MODULE.__dict__["JSON_PATH"] = _JSON_PATH
_V08C_MODULE.__dict__["XLSX_PATH"] = _XLSX_PATH
_V08C_MODULE.__dict__["LLM_JSONL_PATH"] = _LLM_JSONL_PATH
_V08C_MODULE.__dict__["REVIEW_XLSX_PATH"] = _REVIEW_XLSX_PATH
_V08C_MODULE.__dict__["REPORT_PATH"] = _REPORT_PATH
_V08C_MODULE.__dict__["RUN_LABEL"] = _RUN_LABEL
_V08C_MODULE.__dict__["EXTRACTOR_VERSION"] = _EXTRACTOR_VERSION

# 扩展 CANDIDATE_COLUMNS（比 v0.8c 多 extraction_priority 和 index_target_pages）
_CANDIDATE_COLUMNS = [
    "sample_id", "stock_code", "short_name", "report_type",
    "field_id", "dimension", "metric_name_cn", "metric_type", "value_type",
    "indicator_layer", "extraction_priority", "primary_indicator_id",
    "rating_role", "alternative_status_policy", "scoring_denominator_policy",
    "candidate_status", "candidate_disclosure_class", "candidate_rank",
    "evidence_type_candidate", "value_candidate", "unit_raw_candidate",
    "value_standardized_candidate", "unit_standardized_candidate",
    "value_status", "value_extraction_method",
    "source_page", "source_physical_page", "source_report_page_candidates",
    "source_text", "source_table_cell", "match_terms",
    "rule_score", "confidence_rule", "needs_llm_review", "review_reason",
    "recommended_next_status", "extractor_version", "pdf_path",
    "index_target_pages",
]


def load_inputs_v09() -> tuple[list[dict[str, any]], list[dict[str, any]]]:
    """v0.9版加载：支持80指标全量，按PILOT_PRIORITY过滤。"""
    samples = _json.loads(_SAMPLE_JSON.read_text(encoding="utf-8"))["samples"]
    indicators = _json.loads(_INDICATOR_JSON.read_text(encoding="utf-8"))["indicators"]
    sample_map = {item["sample_id"]: item for item in samples}
    selected_ids = choose_sample_ids(samples)
    missing = [sid for sid in selected_ids if sid not in sample_map]
    if missing:
        raise ValueError(f"Unknown sample_id(s): {', '.join(missing)}")
    selected = [sample_map[sid] for sid in selected_ids]

    priority = _EXTRACTION_PRIORITIES.strip()
    if priority and priority.lower() != "all":
        allowed = {p.strip() for p in priority.split(",") if p.strip()}
        filtered = [item for item in indicators if item.get("extraction_priority", "P0") in allowed]
        print(f"抽取优先级过滤: {priority} → {len(filtered)}/{len(indicators)} 个指标")
    else:
        filtered = list(indicators)
        print(f"抽取优先级: all → {len(filtered)} 个指标（全量80）")

    return selected, filtered


def find_candidates_for_indicator_v09(
    sample: dict[str, any],
    pdf_payload: dict[str, any],
    indicator: dict[str, any],
    index_target_map: dict[str, list[int]] | None = None,
) -> list[dict[str, any]]:
    """
    v0.9增强版候选查找：在原有逻辑基础上，利用索引页反向定位结果
    优先检索索引指向的目标页及邻域。
    """
    terms = terms_for_indicator(indicator)
    field_id = indicator["field_id"]

    # ── v0.9新增：索引目标页优先检索 ──
    index_target_pages: set[int] = set()
    if index_target_map and field_id in index_target_map:
        index_target_pages = set(index_target_map[field_id])
        # 扩展邻域（±2页）
        expanded = set(index_target_pages)
        for p in index_target_pages:
            for delta in [-2, -1, 1, 2]:
                expanded.add(p + delta)
        index_target_pages = expanded

    # 原有逻辑：表格候选
    table_candidates: list[tuple[float, dict[str, any], list[str]]] = []
    for row in pdf_payload["table_rows"]:
        row_text = row.get("row_text", "")
        scored_text = f"{row.get('header_context', '')} {row_text}"
        base_score, hits = page_score(row_text, terms)

        # v0.9增强：索引目标页加分
        index_bonus = 0.0
        if index_target_pages and int(row.get("page", 0)) in index_target_pages:
            index_bonus = 20.0

        score = (
            base_score
            + index_bonus
            + evidence_page_adjustment(scored_text, indicator)
            + evidence_structure_bonus(scored_text, indicator)
        )
        if base_score > 0 and score > 0:
            table_candidates.append((score + 8, row, hits))

    # 原有逻辑：文本候选（含索引目标页优先）
    text_candidates: list[tuple[float, dict[str, any], list[str]]] = []
    for page in pdf_payload["pages"]:
        base_score, hits = page_score(page["text"], terms)
        short_label_hits = []
        if base_score <= 0:
            short_label_hits = ocr_short_label_hits(page, indicator)
            if short_label_hits:
                base_score = 18 + len(short_label_hits) * 4
                hits = short_label_hits

        # v0.9增强：索引目标页加分
        index_bonus = 0.0
        if index_target_pages and int(page.get("page", 0)) in index_target_pages:
            index_bonus = 20.0
        # 如果索引指向该页但当前没有直接命中，降低门槛
        if index_target_pages and int(page.get("page", 0)) in index_target_pages and base_score <= 0:
            # 在索引目标页上做一次全量数字扫描作为兜底
            page_text = page.get("text", "")
            number_matches = non_year_number_matches(page_text)
            if number_matches and any(
                term and term in page_text for term in terms
            ):
                base_score = 12.0
                hits = ["index_target_page_fallback"]

        score = (
            base_score
            + index_bonus
            + evidence_page_adjustment(page["text"], indicator)
            + evidence_structure_bonus(page["text"], indicator)
        )
        if base_score > 0 and score > 0:
            text_candidates.append((score, page, hits))

    # ── 构建候选记录（与原逻辑一致，但增加 index_target_pages 字段）──
    records: list[dict[str, any]] = []
    rank = 1
    for score, row, hits in sorted(table_candidates, key=lambda item: item[0], reverse=True)[:2]:
        source = row["row_text"]
        source_pages, physical_page, report_page_candidates = source_page_reference(pdf_payload, row["page"])
        table_context = (
            f"header: {row.get('header_context', '')} || "
            f"table_{row['table_index']}_row_{row['row_index']}: {source}"
        )
        record = build_candidate_record(
            sample, indicator, rank, "candidate_found", "native_table",
            score, source_pages, physical_page, report_page_candidates,
            table_context, table_context, hits, row,
        )
        if keep_candidate(record, indicator):
            # 添加 extraction_priority 和 index_target_pages
            record["extraction_priority"] = indicator.get("extraction_priority", "P0")
            record["index_target_pages"] = ";".join(str(p) for p in sorted(index_target_pages))
            records.append(record)
            rank += 1

    text_limit = 2 if records else 3
    for score, page, hits in sorted(text_candidates, key=lambda item: item[0], reverse=True)[:text_limit]:
        snippet = snippet_around(page["text"], [h.split("~")[0] for h in hits])
        physical_page = str(page["page"])
        report_page_candidates = ";".join(str(item) for item in page.get("report_page_candidates") or [])
        source_pages = ";".join(unique_preserve_order(
            ([physical_page] if physical_page else []) + list(page.get("report_page_candidates") or [])
        ))
        if any(
            existing["source_physical_page"] == physical_page and existing["source_text"] == snippet
            for existing in records
        ):
            continue
        record = build_candidate_record(
            sample, indicator, rank, "candidate_found",
            page.get("text_source") or "native_text",
            score, source_pages, physical_page, report_page_candidates,
            snippet, "", hits, page_payload=page,
        )
        if keep_candidate(record, indicator):
            record["extraction_priority"] = indicator.get("extraction_priority", "P0")
            record["index_target_pages"] = ";".join(str(p) for p in sorted(index_target_pages))
            records.append(record)
            rank += 1

    if not records:
        record = build_candidate_record(
            sample, indicator, 1, "no_candidate", "", 0.0,
            "", "", "", "", "", [],
        )
        record["extraction_priority"] = indicator.get("extraction_priority", "P0")
        record["index_target_pages"] = ";".join(str(p) for p in sorted(index_target_pages))
        records.append(record)
    return records[:3]


def main_v09() -> None:
    _OUT_DIR.mkdir(parents=True, exist_ok=True)
    _PAGE_TEXT_DIR.mkdir(parents=True, exist_ok=True)
    samples, indicators = load_inputs_v09()

    all_terms: list[str] = []
    for indicator in indicators:
        all_terms.extend(split_aliases(
            indicator["metric_name_cn"],
            indicator.get("aliases_cn", ""),
            indicator.get("aliases_en", ""),
        ))
    global_clues = sorted(set(_V08C_MODULE.__dict__["TABLE_CUES"] + all_terms), key=len, reverse=True)

    start = _v9_time.time()
    pdf_payloads: dict[str, dict[str, any]] = {}
    for sample in samples:
        print(f"extracting {sample['sample_id']} {sample['stock_code']} {sample['short_name']}")
        pdf_payloads[sample["sample_id"]] = extract_pdf(sample, global_clues)

    # ── v0.9 新增：索引页反向定位 ──
    indicator_map = {item["field_id"]: item for item in indicators}
    index_target_maps: dict[str, dict[str, list[int]]] = {}
    index_resolver_entries_all: dict[str, list[dict[str, any]]] = {}

    if _ENABLE_INDEX_RESOLVER:
        print("\n=== 索引页反向定位 ===")
        for sample in samples:
            sid = sample["sample_id"]
            payload = pdf_payloads[sid]
            target_map = resolve_index_target_pages(
                payload, sid, indicator_map, _OCR_JSON_DIR, _OCR_TEXT_DIR,
            )
            if target_map:
                index_target_maps[sid] = target_map
                total_pages = sum(len(pages) for pages in target_map.values())
                unique_pages = len(set(p for pages in target_map.values() for p in pages))
                print(f"  {sid} {sample['short_name']}: 索引定位到 {len(target_map)} 个指标的 {unique_pages} 个目标物理页")

    # ── 抽取循环（使用 v0.9 增强版候选查找）──
    records: list[dict[str, any]] = []
    for sample in samples:
        sid = sample["sample_id"]
        payload = pdf_payloads[sid]
        idx_map = index_target_maps.get(sid)
        for indicator in indicators:
            records.extend(find_candidates_for_indicator_v09(sample, payload, indicator, idx_map))

    # ── v2.3: 跨指标去重 ──
    indicator_map = {item["field_id"]: item for item in indicators}
    records = deduplicate_cross_indicator(records, indicator_map)

    summaries = make_summaries(records, samples, indicators)

    # ── 写输出 ──
    # CSV
    with _CSV_PATH.open("w", encoding="utf-8-sig", newline="") as handle:
        import csv as _csv
        writer = _csv.DictWriter(handle, fieldnames=_CANDIDATE_COLUMNS)
        writer.writeheader()
        # 补充 extraction_priority 和 index_target_pages 到已有记录
        for r in records:
            if "extraction_priority" not in r:
                r["extraction_priority"] = indicator_map.get(r.get("field_id", ""), {}).get("extraction_priority", "")
            if "index_target_pages" not in r:
                r["index_target_pages"] = ""
        writer.writerows(records)

    # JSON
    payload_out = {
        "generated_at": _datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "extractor_version": _EXTRACTOR_VERSION,
        "pilot_sample_ids": [s["sample_id"] for s in samples],
        "indicator_count": len(indicators),
        "priority_filter": _EXTRACTION_PRIORITIES,
        "index_resolver_enabled": _ENABLE_INDEX_RESOLVER,
        "samples": samples,
        "indicators": indicators,
        "summaries": summaries,
        "records": records,
    }
    _JSON_PATH.write_text(_json.dumps(payload_out, ensure_ascii=False, indent=2), encoding="utf-8")

    # Excel
    wb = _v9_Workbook()
    ws = wb.active
    ws.title = "候选抽取明细"
    safe_append_rows(ws, records, _CANDIDATE_COLUMNS)

    sample_ws = wb.create_sheet("样本汇总")
    sample_cols = list(summaries["sample_summary"][0].keys()) if summaries["sample_summary"] else []
    safe_append_rows(sample_ws, summaries["sample_summary"], sample_cols)

    indicator_ws = wb.create_sheet("指标汇总")
    indicator_cols = list(summaries["indicator_summary"][0].keys()) if summaries["indicator_summary"] else []
    safe_append_rows(indicator_ws, summaries["indicator_summary"], indicator_cols)

    overall_ws = wb.create_sheet("总体统计")
    overall_ws.append(["指标", "值"])
    for key, value in summaries["overall"].items():
        overall_ws.append([key, _json.dumps(value, ensure_ascii=False) if isinstance(value, dict) else value])
    style_sheet(overall_ws)
    overall_ws.column_dimensions["A"].width = 30
    overall_ws.column_dimensions["B"].width = 90
    wb.save(_XLSX_PATH)

    # LLM队列和复核清单
    write_llm_queue(records, indicators)
    write_review_queue_safe(records)

    elapsed = _v9_time.time() - start

    # ── 写索引解析报告 ──
    if _ENABLE_INDEX_RESOLVER and any(index_target_maps.values()):
        index_report_lines = [
            "# 索引页反向定位报告 v0.9",
            "",
            f"生成时间：{_datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## 概述",
            "",
            "本报告展示从样本 ESG 报告的指标索引页中解析出的章节-页码映射，",
            "以及这些映射与 80 个 ESG 指标的对应关系。",
            "",
            "## 各样本索引解析结果",
            "",
        ]
        for sample in samples:
            sid = sample["sample_id"]
            target_map = index_target_maps.get(sid, {})
            if not target_map:
                index_report_lines.append(f"### {sid} {sample['short_name']}: 未检测到可用索引页")
                continue
            index_report_lines.extend([
                f"### {sid} {sample['short_name']} ({sample['stock_code']})",
                "",
                f"索引定位到 {len(target_map)} 个指标：",
                "",
                "| field_id | 指标名称 | 目标物理页 | 优先层级 |",
                "| --- | --- | --- | --- |",
            ])
            for field_id, pages in sorted(target_map.items()):
                ind = indicator_map.get(field_id, {})
                name = ind.get("metric_name_cn", field_id)
                pages_str = ",".join(str(p) for p in sorted(pages))
                priority = ind.get("extraction_priority", "")
                index_report_lines.append(f"| {field_id} | {name} | {pages_str} | {priority} |")
            index_report_lines.append("")

        _INDEX_REPORT_PATH.write_text("\n".join(index_report_lines), encoding="utf-8")

    # ── 写主报告 ──
    overall = summaries["overall"]
    priority_counts = dict(_Counter(ind.get("extraction_priority", "P0") for ind in indicators))
    report_lines = [
        "# 全量指标候选抽取试跑报告 v0.9",
        "",
        f"生成时间：{_datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "",
        "## 试跑设置",
        "",
        f"- 抽取版本：{_EXTRACTOR_VERSION}",
        f"- 指标范围：{_EXTRACTION_PRIORITIES}，共 {len(indicators)} 个指标",
        f"- 优先级分布：{_json.dumps(priority_counts, ensure_ascii=False)}",
        f"- 索引反向定位：{'启用' if _ENABLE_INDEX_RESOLVER else '未启用'}",
        f"- 样本数：{len(samples)}",
        f"- 技术路线：PDF原生文本+表格+OCR缓存+索引页反向定位 → 候选抽取 → LLM/人工校验",
        f"- 运行耗时：{round(elapsed, 2)} 秒",
        "",
        "## 总体结果",
        "",
        f"- 字段-样本组合：{overall['field_sample_pairs_total']} 个",
        f"- 找到候选证据的组合：{overall['fields_with_candidate_total']} 个",
        f"- 候选覆盖率：{overall['field_sample_candidate_rate']:.2%}",
        f"- 候选明细行：{overall['candidate_records']} 行",
        f"- 候选命中行：{overall['candidate_found_rows']} 行",
        f"- 未命中占位行：{overall['no_candidate_rows']} 行",
        f"- 证据类型分布：{_json.dumps(overall['evidence_type_counts'], ensure_ascii=False)}",
        f"- 披露类别分布：{_json.dumps(overall['candidate_disclosure_class_counts'], ensure_ascii=False)}",
        "",
        "## 样本维度汇总",
        "",
        md_table(
            summaries["sample_summary"],
            ["sample_id", "stock_code", "short_name", "candidate_field_count",
             "p0_field_count", "candidate_field_rate", "table_candidate_rows",
             "text_candidate_rows", "ocr_text_candidate_rows", "no_candidate_fields"],
        ),
        "",
        "## 索引反向定位效果",
        "",
    ]
    if _ENABLE_INDEX_RESOLVER:
        for sample in samples:
            sid = sample["sample_id"]
            tmap = index_target_maps.get(sid, {})
            if tmap:
                report_lines.append(f"- {sid} {sample['short_name']}: {len(tmap)} 个指标有索引定位")
            else:
                report_lines.append(f"- {sid} {sample['short_name']}: 无索引定位（无索引页或已通过其他方式覆盖）")
    else:
        report_lines.append("- 索引反向定位未启用")
    report_lines.extend([
        "",
        "## 产物位置",
        "",
        f"- 候选明细CSV：`{_CSV_PATH}`",
        f"- 候选明细Excel：`{_XLSX_PATH}`",
        f"- 候选明细JSON：`{_JSON_PATH}`",
        f"- LLM校验JSONL：`{_LLM_JSONL_PATH}`",
        f"- 人工复核清单：`{_REVIEW_XLSX_PATH}`",
        f"- 索引解析报告：`{_INDEX_REPORT_PATH}`",
        "",
        "## 下一步",
        "",
        "1. 运行 build_disclosure_matrix_v04.py 生成 v0.4 披露矩阵（涵盖全量80指标）",
        "2. 运行 triage_not_found_review_v02.py 生成 v0.2 分流",
        "3. 运行 build_scoring_input_v02.py 生成 v0.2 评分输入",
        "4. 运行 build_static_dashboard_v02.py 生成 v0.2 仪表盘",
        "5. 对 GL020 等索引样本做专项回归验证",
    ])
    _REPORT_PATH.write_text("\n".join(report_lines), encoding="utf-8")

    print(_json.dumps({
        "elapsed_sec": round(elapsed, 2),
        "indicator_count": len(indicators),
        "priority_filter": _EXTRACTION_PRIORITIES,
        "index_resolver_enabled": _ENABLE_INDEX_RESOLVER,
        "samples_with_index": len([s for s in samples if index_target_maps.get(s["sample_id"])]),
        **summaries["overall"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main_v09()
